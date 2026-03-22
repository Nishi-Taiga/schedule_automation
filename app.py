#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Booth Schedule Generator – Cloud Edition (Render)
Flask + gunicorn + openpyxl
"""
import os
import sys
import json
import random
import re
import threading
import tempfile
import shutil
import time
import secrets
import atexit
import traceback
import datetime as _dt
import calendar
import base64
import zipfile
import io
import hashlib
import gzip
from copy import copy, deepcopy
from collections import defaultdict
from functools import wraps
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB上限
# SECRET_KEY: 環境変数推奨。未設定時はランダム生成（再起動でセッション無効化）
_secret_key = os.environ.get('SECRET_KEY', '')
if not _secret_key:
    _secret_key = secrets.token_hex(32)
    print("[SECURITY WARNING] SECRET_KEY が未設定です。ランダムキーを生成しました。"
          "再起動でセッションが無効化されます。本番環境では環境変数に設定してください。", flush=True)
app.secret_key = _secret_key

# パスワード（環境変数必須。未設定時はランダム生成＋ログ表示）
APP_PASSWORD = os.environ.get('APP_PASSWORD', '')
if not APP_PASSWORD:
    APP_PASSWORD = secrets.token_urlsafe(12)
    print(f"[SECURITY WARNING] APP_PASSWORD が未設定です。一時パスワード: {APP_PASSWORD}", flush=True)

# ========== セキュリティヘッダー ==========
@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'camera=(), microphone=(), geolocation=()'
    return response

# 一時ファイル管理（ディスクベース: gunicornマルチワーカー対応）
UPLOAD_BASE = os.path.join(tempfile.gettempdir(), 'booth_sessions')
os.makedirs(UPLOAD_BASE, exist_ok=True)
SESSION_TIMEOUT = 3600  # 1時間でクリーンアップ

def _session_dir(sid):
    """セッションIDからディレクトリパスを返す"""
    return os.path.join(UPLOAD_BASE, sid)

def _session_meta_path(sid):
    """セッションのメタデータJSONファイルパス"""
    return os.path.join(_session_dir(sid), '_meta.json')

def _load_meta(sid):
    """ディスクからセッションメタデータを読み込む"""
    mp = _session_meta_path(sid)
    if os.path.exists(mp):
        with open(mp, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None

def _save_meta(sid, meta):
    """セッションメタデータをディスクに保存"""
    mp = _session_meta_path(sid)
    with open(mp, 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False)

def cleanup_old_sessions():
    """古いセッションの一時ファイルを削除"""
    now = time.time()
    if not os.path.exists(UPLOAD_BASE):
        return
    for name in os.listdir(UPLOAD_BASE):
        sdir = os.path.join(UPLOAD_BASE, name)
        if not os.path.isdir(sdir):
            continue
        meta = _load_meta(name)
        if meta and now - meta.get('last_access', 0) > SESSION_TIMEOUT:
            shutil.rmtree(sdir, ignore_errors=True)
    # Supabase: 7日以上古いセッションを削除
    cutoff = (_dt.datetime.utcnow() - _dt.timedelta(days=7)).isoformat() + 'Z'
    _supabase_request('DELETE', 'schedule_sessions', f'updated_at=lt.{cutoff}')

def get_session_data():
    """現在のセッションのデータを取得(なければ作成) - ディスクベース"""
    cleanup_old_sessions()
    sid = session.get('sid')
    sdir = _session_dir(sid) if sid else None
    if not sid or not os.path.exists(_session_meta_path(sid)):
        sid = secrets.token_hex(16)
        session['sid'] = sid
        sdir = _session_dir(sid)
        os.makedirs(sdir, exist_ok=True)
        meta = {'files': {}, 'dir': sdir, 'last_access': time.time()}
        _save_meta(sid, meta)
    meta = _load_meta(sid)
    meta['last_access'] = time.time()
    meta['dir'] = sdir  # 常にパスを保証
    _save_meta(sid, meta)
    # resultはインメモリで保持（大きいため）、ただしfilesパスはディスクから復元
    if not hasattr(get_session_data, '_cache'):
        get_session_data._cache = {}
    if sid not in get_session_data._cache:
        get_session_data._cache[sid] = {'result': {}}
    cached = get_session_data._cache[sid]
    return {**meta, 'result': cached.get('result', {}), '_sid': sid}

def save_session_files(sd):
    """ファイルパス情報をディスクに保存"""
    sid = sd['_sid']
    meta = _load_meta(sid)
    meta['files'] = sd['files']
    meta['last_access'] = time.time()
    _save_meta(sid, meta)

def save_session_result(sd):
    """resultをインメモリキャッシュ + ディスクに保存"""
    sid = sd['_sid']
    if not hasattr(get_session_data, '_cache'):
        get_session_data._cache = {}
    get_session_data._cache[sid] = {'result': sd['result']}
    # ディスクにもJSON保存（サーバー再起動後の復元用）
    _save_result_to_disk(sid, sd['result'])

def _result_json_path(sid):
    return os.path.join(_session_dir(sid), '_result.json')

def _save_result_to_disk(sid, result):
    """スケジュール結果をディスクにJSON保存"""
    rp = _result_json_path(sid)
    try:
        # schedule内のtupleをlistに変換して保存
        saveable = {}
        if 'schedule_json' in result:
            saveable['schedule_json'] = result['schedule_json']
        if 'original_schedule_json' in result:
            saveable['original_schedule_json'] = result['original_schedule_json']
        if 'original_unplaced' in result:
            saveable['original_unplaced'] = result['original_unplaced']
        if 'unplaced' in result:
            saveable['unplaced'] = result['unplaced']
        if 'office_teachers' in result:
            saveable['office_teachers'] = result['office_teachers']
        if 'booth_pref' in result:
            saveable['booth_pref'] = result['booth_pref']
        if 'students' in result:
            # studentsのsetをlistに変換
            stu_save = []
            for s in result['students']:
                sc = dict(s)
                if isinstance(sc.get('avail'), set):
                    sc['avail'] = sorted([list(a) for a in sc['avail']])
                if isinstance(sc.get('backup_avail'), set):
                    sc['backup_avail'] = sorted([list(a) for a in sc['backup_avail']])
                if isinstance(sc.get('ng_dates'), set):
                    sc['ng_dates'] = [list(d) for d in sc['ng_dates']]
                if 'fixed' in sc:
                    sc['fixed'] = [list(f) for f in sc['fixed']]
                stu_save.append(sc)
            saveable['students'] = stu_save
        if 'week_dates' in result:
            saveable['week_dates'] = result['week_dates']
        if 'weekly_teachers' in result:
            saveable['weekly_teachers'] = result['weekly_teachers']
        if 'skills' in result:
            saveable['skills'] = result['skills']
        with open(rp, 'w', encoding='utf-8') as f:
            json.dump(saveable, f, ensure_ascii=False)
        # Supabaseにも永続保存
        _save_result_to_supabase(sid, saveable)
    except Exception as e:
        print(f"[save_result] WARNING: ディスク保存失敗: {e}", flush=True)

def _save_result_to_supabase(sid, saveable):
    """スケジュール結果をSupabaseに永続保存 (upsert)"""
    try:
        sid = _sanitize_postgrest_value(sid, 'sid')
        _supabase_request('POST', 'schedule_sessions', '', body={
            'sid': sid,
            'result_data': saveable,
            'updated_at': _dt.datetime.utcnow().isoformat() + 'Z',
        }, headers_extra={'Prefer': 'resolution=merge-duplicates'})
    except Exception as e:
        print(f"[save_result] WARNING: Supabase保存失敗: {e}", flush=True)

def _load_result_from_disk(sid):
    """ディスクからスケジュール結果を読み込む"""
    rp = _result_json_path(sid)
    if not os.path.exists(rp):
        return None
    try:
        with open(rp, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None

def _load_result_from_supabase(sid):
    """Supabaseからスケジュール結果を読み込む"""
    try:
        sid = _sanitize_postgrest_value(sid, 'sid')
        rows = _supabase_request('GET', 'schedule_sessions', f'sid=eq.{sid}&select=result_data')
        if rows and len(rows) > 0:
            return rows[0].get('result_data')
    except Exception as e:
        print(f"[load_result] WARNING: Supabase読み込み失敗: {e}", flush=True)
    return None

# ========== 認証 ==========
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('authenticated'):
            if request.is_json or request.path.startswith('/api/'):
                return jsonify({'error': '認証が必要です'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

# ========== ログインレートリミッター ==========
_login_attempts = {}  # {ip: [(timestamp, ...), ...]}
_login_lock = threading.Lock()
_LOGIN_MAX_ATTEMPTS = 5
_LOGIN_WINDOW_SEC = 60
_LOGIN_LOCKOUT_SEC = 300

def _check_login_rate_limit(ip):
    """ログイン試行のレートリミット。ロックアウト中なら残り秒数を返す、OKなら0"""
    now = time.time()
    with _login_lock:
        attempts = _login_attempts.get(ip, [])
        # 古いエントリを除去
        attempts = [t for t in attempts if now - t < _LOGIN_LOCKOUT_SEC]
        _login_attempts[ip] = attempts
        # ウィンドウ内の試行回数
        recent = [t for t in attempts if now - t < _LOGIN_WINDOW_SEC]
        if len(recent) >= _LOGIN_MAX_ATTEMPTS:
            oldest = min(recent)
            return int(_LOGIN_LOCKOUT_SEC - (now - oldest))
    return 0

def _record_login_failure(ip):
    """ログイン失敗を記録"""
    with _login_lock:
        _login_attempts.setdefault(ip, []).append(time.time())

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    error = None
    if request.method == 'POST':
        ip = request.remote_addr or 'unknown'
        lockout = _check_login_rate_limit(ip)
        if lockout > 0:
            error = f'ログイン試行回数を超えました。{lockout}秒後に再試行してください'
            return render_template('login.html', error=error)
        pw = request.form.get('password', '')
        if pw == APP_PASSWORD:
            session['authenticated'] = True
            return redirect(url_for('index'))
        _record_login_failure(ip)
        error = 'パスワードが違います'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    sid = session.get('sid')
    if sid:
        sdir = _session_dir(sid)
        if os.path.exists(sdir):
            shutil.rmtree(sdir, ignore_errors=True)
        if hasattr(get_session_data, '_cache') and sid in get_session_data._cache:
            del get_session_data._cache[sid]
        try:
            safe_sid = _sanitize_postgrest_value(sid, 'sid')
            _supabase_request('DELETE', 'schedule_sessions', f'sid=eq.{safe_sid}')
        except ValueError:
            pass  # 不正なsidの場合はSupabase削除をスキップ
    session.clear()
    return redirect(url_for('login_page'))

ALLOWED_EXT = {'.xlsx'}
def validate_file(f):
    if not f or not f.filename:
        return False, 'ファイルが選択されていません'
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXT:
        return False, f'許可されていないファイル形式です: {ext}'
    return True, None

# ========== 定数 ==========
DAYS = ['月','火','水','木','金','土']
WEEKDAY_TIMES = ['16:00','17:05','18:10','19:15','20:20']
SATURDAY_TIMES = ['14:55','16:00','17:05','18:10']
ALL_TIMES = ['14:55','16:00','17:05','18:10','19:15','20:20']
TIME_SHORT = {'14:55':'14','16:00':'16','17:05':'17','18:10':'18','19:15':'19','20:20':'20'}
TIME_SHORT_REV = {v: k for k, v in TIME_SHORT.items()}
MAX_BOOTHS = 6
# メタシート判定キーワード（週シートと区別するため）
# load_teacher_skills の検出キーワードと一致させること
META_KEYWORDS = ['必要コマ', '一覧', 'ブース希望', '指導可能', 'スキル']

NAME_MAP = {}  # 動的に構築（_build_name_map で同姓講師を自動検出）

def _build_name_map(full_names):
    """同姓講師を自動検出し、名前（ファーストネーム）+'T' で区別する。"""
    global NAME_MAP
    NAME_MAP.clear()
    surname_groups = defaultdict(list)
    seen = set()
    for full in full_names:
        full_str = str(full).strip()
        if full_str in seen:
            continue
        seen.add(full_str)
        parts = full_str.replace('\u3000', ' ').split()
        if len(parts) >= 2:
            surname_groups[parts[0]].append((full_str, parts))
    for surname, entries in surname_groups.items():
        if len(entries) > 1:
            for full, parts in entries:
                NAME_MAP[full] = parts[1] + 'T'

# デフォルト講師ブース希望（ブース表から読み込み or UIで設定）
DEFAULT_BOOTH_PREF = {}

SRC_TIME_SLOTS = [
    (6,'14:55',6),(19,'16:00',6),(32,'17:05',6),
    (45,'18:10',6),(58,'19:15',9),(77,'20:20',9),
]
SRC_DAY_COLS = {'月':3,'火':8,'水':13,'木':18,'金':23,'土':28}

SKILL_COL_MAP = {
    3:'小国',4:'小算',5:'小英',6:'小理',7:'小社',
    8:'受国',9:'受算',10:'受理',11:'受社',
    12:'中国',13:'中数',14:'中英',15:'中理',16:'中社',
    17:'高現',18:'高古',
}

LAYOUT = {
    '14:55':(7,6),'16:00':(20,6),'17:05':(33,6),
    '18:10':(46,6),'19:15':(59,6),'20:20':(72,6),
}
DAY_COLS = {
    '月':(3,4,5,6,7),'火':(8,9,10,11,12),'水':(13,14,15,16,17),
    '木':(18,19,20,21,22),'金':(23,24,25,26,27),'土':(28,29,30,31,32),
}
TUTOR_ROWS = [19,32,45,58,71,84]

# ========== 学習システム ==========
SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_KEY', '')

DEFAULT_WEIGHTS = {
    'ng_date': -5000,
    'backup_time': -2100,
    'continuous_block': 2000,
    'skip_interval': -200,
    'same_day_2nd': 50,
    'same_day_3plus': -80,
    'wish_teacher': 500,
    'booth_pref': 10,
    'empty_booth': 20,
}

WEIGHT_BOUNDS = {
    'ng_date': (-10000, -1000),
    'backup_time': (-3000, 0),
    'continuous_block': (500, 5000),
    'skip_interval': (-1000, 0),
    'same_day_2nd': (-50, 200),
    'same_day_3plus': (-300, 50),
    'wish_teacher': (100, 2000),
    'booth_pref': (0, 50),
    'empty_booth': (0, 100),
}

def _sanitize_postgrest_value(value, expected_type='string'):
    """PostgREST クエリパラメータのバリデーション。不正値は ValueError を送出"""
    if value is None:
        raise ValueError('値が必要です')
    s = str(value).strip()
    if not s:
        raise ValueError('空の値は許可されません')
    if expected_type == 'uuid':
        if not re.match(r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$', s):
            raise ValueError('不正なID形式です')
        return s
    if expected_type == 'int':
        try:
            n = int(s)
        except (ValueError, TypeError):
            raise ValueError('整数値が必要です')
        if n < 0 or n > 9999:
            raise ValueError('数値が範囲外です')
        return n
    if expected_type == 'label':
        if len(s) > 50:
            raise ValueError('ラベルが長すぎます')
        if not re.match(r'^[\w\u3000-\u9FFF\u30A0-\u30FF\u3040-\u309F\-]+$', s):
            raise ValueError('ラベルに不正な文字が含まれています')
        return s
    if expected_type == 'sid':
        if not re.match(r'^[a-zA-Z0-9_\-]{8,64}$', s):
            raise ValueError('不正なセッションID形式です')
        return s
    # default: 危険文字を拒否 (PostgREST演算子注入防止)
    if any(c in s for c in '&=|;'):
        raise ValueError('不正な文字が含まれています')
    return s


def _supabase_request(method, table, params='', body=None, headers_extra=None):
    """Supabase REST API へのリクエストヘルパー"""
    if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
        return None
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += f"?{params}"
    hdrs = {
        'apikey': SUPABASE_SERVICE_KEY,
        'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
        'Content-Type': 'application/json',
    }
    if headers_extra:
        hdrs.update(headers_extra)
    data = json.dumps(body, ensure_ascii=False).encode('utf-8') if body else None
    req = Request(url, data=data, headers=hdrs, method=method)
    try:
        with urlopen(req, timeout=10) as resp:
            raw = resp.read().decode('utf-8')
            return json.loads(raw) if raw.strip() else None
    except (URLError, HTTPError) as e:
        print(f"[learning] Supabase {method} {table} error: {e}", flush=True)
        return None

def _encode_booth_files(sd):
    """セッションのブース表ファイル群(meta+week_files)をZIP→base64エンコード。b64_str or None"""
    files = sd.get('files', {})
    booth_path = files.get('booth')
    week_file_paths = files.get('week_files', [])

    # メタもウィークもなければスキップ
    if (not booth_path or not os.path.exists(booth_path)) and not week_file_paths:
        return None

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        if booth_path and os.path.exists(booth_path):
            zf.write(booth_path, 'meta/' + os.path.basename(booth_path))
        for wp in week_file_paths:
            if os.path.exists(wp):
                zf.write(wp, 'weeks/' + os.path.basename(wp))
    data = buf.getvalue()
    if len(data) > 10 * 1024 * 1024:  # 10MB上限
        print(f"[cloud_save] booth zip too large: {len(data)} bytes, skipping", flush=True)
        return None
    print(f"[cloud_save] booth zip size: {len(data)} bytes (meta={'yes' if booth_path else 'no'}, weeks={len(week_file_paths)})", flush=True)
    return base64.b64encode(data).decode('ascii')

def _restore_booth_files(b64_str, session_dir):
    """base64 ZIPからブース表ファイル群を復元。{'booth': path, 'week_files': [paths]} or None"""
    if not b64_str:
        return None
    try:
        raw = base64.b64decode(b64_str)
        buf = io.BytesIO(raw)
        result = {}
        real_session_dir = os.path.realpath(session_dir)
        with zipfile.ZipFile(buf, 'r') as zf:
            for name in zf.namelist():
                dest = os.path.realpath(os.path.join(session_dir, name.replace('/', os.sep)))
                # パストラバーサル防止: セッションディレクトリ外への展開をブロック
                if not dest.startswith(real_session_dir + os.sep) and dest != real_session_dir:
                    print(f"[cloud_load] path traversal blocked: {name}", flush=True)
                    continue
                os.makedirs(os.path.dirname(dest), exist_ok=True)
                with zf.open(name) as src, open(dest, 'wb') as dst:
                    dst.write(src.read())
                if name.startswith('meta/'):
                    result['booth'] = dest
                elif name.startswith('weeks/'):
                    result.setdefault('week_files', []).append(dest)
        # week_filesをソート
        if 'week_files' in result:
            result['week_files'] = sorted(result['week_files'])
        print(f"[cloud_load] booth files restored: meta={'booth' in result}, weeks={len(result.get('week_files', []))}", flush=True)
        return result if result else None
    except Exception as e:
        print(f"[cloud_load] booth restore failed: {e}", flush=True)
        return None

def load_learning_weights():
    """Supabaseから学習済み重みを読み込む。なければデフォルト値を返す"""
    rows = _supabase_request('GET', 'schedule_learning_data', 'key=eq.weights&select=data')
    if rows and len(rows) > 0:
        saved = rows[0].get('data', {})
        weights = dict(DEFAULT_WEIGHTS)
        # statsのsession_countを確認（3未満なら学習適用しない）
        stats_rows = _supabase_request('GET', 'schedule_learning_data', 'key=eq.stats&select=data')
        session_count = 0
        if stats_rows and len(stats_rows) > 0:
            session_count = stats_rows[0].get('data', {}).get('session_count', 0)
        if session_count >= 3:
            for k in DEFAULT_WEIGHTS:
                if k in saved:
                    weights[k] = int(round(saved[k]))
        return weights
    return dict(DEFAULT_WEIGHTS)

def save_learning_weights(weights):
    """学習済み重みをSupabaseに保存 (upsert)"""
    _supabase_request('POST', 'schedule_learning_data', '', body={
        'key': 'weights',
        'data': weights,
        'updated_at': _dt.datetime.utcnow().isoformat() + 'Z',
    }, headers_extra={'Prefer': 'resolution=merge-duplicates'})

def load_learning_stats():
    """学習統計を読み込む"""
    rows = _supabase_request('GET', 'schedule_learning_data', 'key=eq.stats&select=data')
    if rows and len(rows) > 0:
        return rows[0].get('data', {})
    return {'session_count': 0}

def save_learning_stats(stats):
    """学習統計をSupabaseに保存 (upsert)"""
    _supabase_request('POST', 'schedule_learning_data', '', body={
        'key': 'stats',
        'data': stats,
        'updated_at': _dt.datetime.utcnow().isoformat() + 'Z',
    }, headers_extra={'Prefer': 'resolution=merge-duplicates'})

def save_edit_history(entry):
    """編集履歴を保存し、20件超を削除"""
    _supabase_request('POST', 'schedule_edit_history', '', body=entry,
                      headers_extra={'Prefer': 'return=minimal'})
    # 古いレコードを削除（最新20件以外）
    rows = _supabase_request('GET', 'schedule_edit_history',
                             'select=id&order=created_at.desc&offset=20')
    if rows:
        for r in rows:
            _supabase_request('DELETE', 'schedule_edit_history', f"id=eq.{r['id']}")

def _index_placements(schedule_json):
    """スケジュールから (name, subject) → [(wi, day, ts, bi, teacher), ...] のインデックスを構築"""
    idx = defaultdict(list)
    for wi, week in enumerate(schedule_json):
        for day, day_data in week.items():
            for ts, booths in day_data.items():
                for bi, b in enumerate(booths):
                    teacher = b.get('teacher', '')
                    for slot in b.get('slots', []):
                        if len(slot) >= 3:
                            name, subj = slot[1], slot[2]
                            idx[(name, subj)].append({
                                'wi': wi, 'day': day, 'ts': ts, 'bi': bi, 'teacher': teacher
                            })
    return idx

def compute_schedule_diff(original, edited, orig_unplaced, edit_unplaced):
    """自動生成スケジュールと手動編集後スケジュールの差分を計算"""
    changes = []
    orig_idx = _index_placements(original)
    edit_idx = _index_placements(edited)
    all_keys = set(orig_idx.keys()) | set(edit_idx.keys())

    for key in all_keys:
        orig_locs = orig_idx.get(key, [])
        edit_locs = edit_idx.get(key, [])
        # 位置をタプル化して比較
        orig_set = {(l['wi'], l['day'], l['ts'], l['bi']) for l in orig_locs}
        edit_set = {(l['wi'], l['day'], l['ts'], l['bi']) for l in edit_locs}
        removed = orig_set - edit_set
        added = edit_set - orig_set

        # 同じブースで講師変更を検出
        for loc in orig_locs:
            pos = (loc['wi'], loc['day'], loc['ts'], loc['bi'])
            if pos in edit_set:
                # 同じ位置にあるが講師が違う場合
                for eloc in edit_locs:
                    epos = (eloc['wi'], eloc['day'], eloc['ts'], eloc['bi'])
                    if epos == pos and eloc['teacher'] != loc['teacher']:
                        changes.append({
                            'type': 'teacher_swap',
                            'student': key[0], 'subject': key[1],
                            'wi': loc['wi'], 'day': loc['day'], 'ts': loc['ts'], 'bi': loc['bi'],
                            'from_teacher': loc['teacher'], 'to_teacher': eloc['teacher'],
                        })

        # 移動の対応付け（removedとaddedを対にする）
        removed_list = list(removed)
        added_list = list(added)
        moved_count = min(len(removed_list), len(added_list))
        for i in range(moved_count):
            r = removed_list[i]
            a = added_list[i]
            r_teacher = next((l['teacher'] for l in orig_locs
                              if (l['wi'], l['day'], l['ts'], l['bi']) == r), '')
            a_teacher = next((l['teacher'] for l in edit_locs
                              if (l['wi'], l['day'], l['ts'], l['bi']) == a), '')
            changes.append({
                'type': 'student_moved',
                'student': key[0], 'subject': key[1],
                'from': {'wi': r[0], 'day': r[1], 'ts': r[2], 'bi': r[3], 'teacher': r_teacher},
                'to': {'wi': a[0], 'day': a[1], 'ts': a[2], 'bi': a[3], 'teacher': a_teacher},
            })

        # 残りのremoved = 配置→未配置
        for i in range(moved_count, len(removed_list)):
            r = removed_list[i]
            changes.append({
                'type': 'student_removed',
                'student': key[0], 'subject': key[1],
                'from': {'wi': r[0], 'day': r[1], 'ts': r[2], 'bi': r[3]},
            })

        # 残りのadded = 未配置→配置
        for i in range(moved_count, len(added_list)):
            a = added_list[i]
            a_teacher = next((l['teacher'] for l in edit_locs
                              if (l['wi'], l['day'], l['ts'], l['bi']) == a), '')
            changes.append({
                'type': 'student_placed',
                'student': key[0], 'subject': key[1],
                'to': {'wi': a[0], 'day': a[1], 'ts': a[2], 'bi': a[3], 'teacher': a_teacher},
            })

    return changes

def extract_signals(changes, original, edited):
    """差分変更からスコアリング重み調整用のシグナルを抽出"""
    signals = {}
    if not changes:
        return signals

    # カウンター
    continuous_kept = 0
    continuous_broken = 0
    backup_to_primary = 0
    backup_kept = 0
    wish_kept = 0
    wish_overridden = 0
    empty_booth_used = 0
    same_day_3plus_created = 0
    same_day_3plus_broken = 0

    for ch in changes:
        if ch['type'] == 'student_moved':
            frm = ch['from']
            to = ch['to']
            # 時間帯の変化を分析
            if frm['ts'] != to['ts']:
                # 連続コマの変化
                try:
                    times = SATURDAY_TIMES if to['day'] == '土' else WEEKDAY_TIMES
                    from_long = TIME_SHORT_REV.get(frm['ts'])
                    to_long = TIME_SHORT_REV.get(to['ts'])
                    if from_long in times and to_long in times:
                        from_idx = times.index(from_long)
                        to_idx = times.index(to_long)
                        if abs(from_idx - to_idx) == 1:
                            continuous_kept += 1
                        elif abs(from_idx - to_idx) > 1:
                            continuous_broken += 1
                except (ValueError, KeyError):
                    pass

        elif ch['type'] == 'teacher_swap':
            wish_overridden += 1

        elif ch['type'] == 'student_placed':
            # 未配置→配置（アルゴリズムが見つけられなかったスロットを人間が発見）
            to = ch['to']
            # 空きブースかどうか確認
            try:
                booth = edited[to['wi']][to['day']][to['ts']][to['bi']]
                if len(booth.get('slots', [])) <= 1:
                    empty_booth_used += 1
            except (IndexError, KeyError):
                pass

    total = len(changes)
    if total == 0:
        return signals

    # 連続コマシグナル
    if continuous_broken + continuous_kept > 0:
        ratio = (continuous_kept - continuous_broken) / (continuous_broken + continuous_kept)
        signals['continuous_block'] = ratio * 0.5

    # 希望講師シグナル
    if wish_overridden > 0:
        signals['wish_teacher'] = -0.3 * min(wish_overridden / total, 1.0)

    # 空きブースシグナル
    if empty_booth_used > 0:
        signals['empty_booth'] = 0.3 * min(empty_booth_used / total, 1.0)

    return signals

def adjust_weights(current_weights, signals, alpha=0.3):
    """EMAで重みを調整する"""
    new_weights = dict(current_weights)
    for key, signal in signals.items():
        if key not in new_weights:
            continue
        delta = signal * abs(DEFAULT_WEIGHTS[key]) * 0.1
        new_weights[key] = new_weights[key] + alpha * delta
        lo, hi = WEIGHT_BOUNDS[key]
        new_weights[key] = max(lo, min(hi, new_weights[key]))
        new_weights[key] = int(round(new_weights[key]))
    return new_weights

def to_short(name):
    if not name: return None
    name = str(name).strip()
    if not name: return None
    if name in NAME_MAP: return NAME_MAP[name]
    parts = name.replace('\u3000',' ').split()
    if len(parts) >= 2:
        return parts[0] + 'T'
    # 既存の破損データ("TT")を修復
    while len(name) > 1 and name.endswith('TT'):
        name = name[:-1]
    # 単一パート: 既に 'T' 末尾なら短縮済み（二重付与を防止）
    if name.endswith('T'):
        return name
    return name + 'T'

def _sanitize_weekly_teachers(wt):
    """weeklyTeachers内の全講師名から重複末尾Tを除去する"""
    if not wt:
        return wt
    result = []
    for week in wt:
        w = {}
        for day, day_data in week.items():
            d = {}
            for ts, teachers in day_data.items():
                d[ts] = [to_short(t) for t in (teachers or [])]
            w[day] = d
        result.append(w)
    return result

# ========== パーサー ==========
def get_skill_keys(grade, subject):
    g = str(grade).upper()
    s = str(subject)
    # 英検は英語スキルでチェック
    if s == '英検':
        s = '英'
    if g.startswith('S'):
        yr = int(g[1:]) if len(g)>1 else 0
        if s == '数': s = '算'  # 小学/受験は「算」
        if yr >= 4:
            # 中学受験には英語がないので小学英語で代替
            if s == '英': return ['小英']
            return ['受'+s]
        return ['小'+s]
    elif g.startswith('C'):
        if s == '算': s = '数'  # 中学は「数」
        return ['中'+s]
    elif g.startswith('K'):
        if s == '算': s = '数'
        if s == '数': return ['高ⅠA','高ⅡB','高Ⅲ','高C']
        return ['高'+s]
    if s == '算': s = '数'
    return ['中'+s]

def can_teach(teacher, grade, subject, skills):
    keys = get_skill_keys(grade, subject)
    if not skills:
        return True
    if teacher not in skills:
        return False
        
    # 英検対応: 「英」を含むスキル（小英、中英、高英など）があれば可
    if subject == '英検':
        return any('英' in k for k in skills[teacher])
        
    return any(k in skills[teacher] for k in keys)

def load_teacher_skills(wb):
    """ブース表xlsx内の講師指導可能科目シートを読み込む"""
    # シート名を自動検出（「一覧表」「指導可能」等を含むシート）
    skill_sheet = None
    for sn in wb.sheetnames:
        if '一覧' in sn or '指導可能' in sn or 'スキル' in sn or 'skill' in sn.lower():
            skill_sheet = sn
            break
    if not skill_sheet:
        return {}  # シートが見つからない場合は空（全講師が全科目可として動作）

    ws = wb[skill_sheet]
    skills = {}
    for r in range(4, ws.max_row+1):
        t = ws.cell(r,2).value
        if not t: break
        t = str(t).strip()
        s = set()
        for c, k in SKILL_COL_MAP.items():
            if ws.cell(r,c).value == '◯': s.add(k)
        for c in range(19, ws.max_column+1):
            v, h = ws.cell(r,c).value, ws.cell(3,c).value
            if v == '◯' and h: s.add('高'+str(h))
        skills[t] = s
    return skills

def load_booth_pref(wb):
    """ブース表xlsx内の講師ブース希望シートを読み込む"""
    for sn in wb.sheetnames:
        if 'ブース希望' in sn:
            ws = wb[sn]
            pref = {}
            for r in range(2, ws.max_row+1):
                t = ws.cell(r, 1).value
                b = ws.cell(r, 2).value
                if t and b:
                    pref[str(t).strip()] = int(b)
            return pref
    return {}

def parse_ng_dates(val, year, month):
    """NG日程を解析して (week_index, day_name) のsetを返す。
    形式例: '2/5', '2/1-2/7', '2/19,2/24,2/25', '12/5', '平日'
    """
    if not val: return set()
    day_names = ['月','火','水','木','金','土','日']
    WEEKDAYS = ['月','火','水','木','金']
    result = set()
    val_str = str(val).strip()

    # 「平日」→ 全週の月〜金をNG
    if '平日' in val_str:
        for wi in range(6):
            for d in WEEKDAYS:
                result.add((wi, d))
        val_str = val_str.replace('平日', '')

    def add_date(m, d):
        if m != month: return
        try:
            dt = _dt.date(year, m, d)
        except ValueError:
            return
        wd = dt.weekday()
        if wd >= 6: return  # 日曜スキップ
        wi = (d - 1) // 7
        result.add((wi, day_names[wd]))

    def parse_md(s):
        s = s.strip().replace('/', '/')
        if '/' in s:
            parts = s.split('/')
            return int(parts[0]), int(parts[1])
        else:
            # 日のみ → 当月と仮定
            return month, int(s)

    for part in val_str.split(','):
        part = part.strip()
        if not part: continue
        if '-' in part:
            # 範囲: 2/1-2/7
            a, b = part.split('-', 1)
            try:
                m1, d1 = parse_md(a)
                m2, d2 = parse_md(b)
                if m1 == m2:
                    for d in range(d1, d2 + 1):
                        add_date(m1, d)
            except (ValueError, IndexError):
                pass
        else:
            try:
                m, d = parse_md(part)
                add_date(m, d)
            except (ValueError, IndexError):
                pass
    return result

def load_students_from_wb(wb, year=2026, month=2):
    ws = wb['必要コマ数']
    subj_cols = [(5,'英'),(6,'英検'),(7,'数'),(8,'算'),(9,'国'),(10,'理'),
                 (11,'社'),(12,'現'),(13,'古'),(14,'物'),(15,'化'),(16,'生'),
                 (17,'日'),(18,'地'),(19,'政'),(20,'世')]
    students = []
    for r in range(3, 60):
        school, grade, name = ws.cell(r,2).value, ws.cell(r,3).value, ws.cell(r,4).value
        if not name: break
        needs = {}
        for col, subj in subj_cols:
            v = ws.cell(r,col).value
            if v and isinstance(v,(int,float)) and v>0: needs[subj] = int(v)
        parse_list = lambda v: [t.strip() for t in str(v or '').split(',') if t.strip()]
        students.append({
            'school':str(school or ''),'grade':str(grade),'name':str(name),'needs':needs,
            'wish_teachers':parse_list(ws.cell(r,21).value),
            'ng_teachers':parse_list(ws.cell(r,22).value),
            'ng_students':parse_list(ws.cell(r,23).value),
            'avail':parse_avail(ws.cell(r,24).value),
            'backup_avail':parse_avail(ws.cell(r,25).value),
            'ng_dates':parse_ng_dates(ws.cell(r,26).value, year, month),
            'fixed':parse_regular(ws.cell(r,27).value),
            'notes':str(ws.cell(r,28).value or '').strip(),
        })
    return students

def parse_avail(val):
    if not val: return None
    WEEKDAYS = ['月','火','水','木','金']
    VALID_DAYS = set(WEEKDAYS + ['土'])
    slots = set()
    for p in str(val).split(','):
        p = p.strip()
        if not p: continue
        # 「平日XX」「平日XX-YY」→ 月〜金に展開
        if p.startswith('平日'):
            rest = p[2:]
            if not rest: continue
            try:
                if '-' in rest:
                    a, b = rest.split('-')
                    for d in WEEKDAYS:
                        for t in range(int(a), int(b)+1): slots.add((d, str(t)))
                else:
                    for d in WEEKDAYS:
                        slots.add((d, rest))
            except (ValueError, IndexError):
                pass
            continue
        # 先頭1文字が有効な曜日でない場合はスキップ（日付や自由テキスト等）
        d, rest = p[0], p[1:]
        if d not in VALID_DAYS:
            continue
        if not rest: continue
        try:
            if '-' in rest:
                a, b = rest.split('-')
                for t in range(int(a), int(b)+1): slots.add((d, str(t)))
            else:
                int(rest)  # 数値チェック
                slots.add((d, rest))
        except (ValueError, IndexError):
            pass
    return slots if slots else None

def parse_regular(val):
    if not val: return []
    WEEKDAYS = ['月','火','水','木','金']
    result = []
    for p in str(val).split(','):
        p = p.strip()
        if ':' not in p: continue
        dt, subj = p.split(':', 1)
        if dt.startswith('平日'):
            ts = dt[2:]
            for d in WEEKDAYS:
                result.append((d, ts, subj.strip()))
        else:
            result.append((dt[0], dt[1:], subj.strip()))
    return result

def load_weekly_teachers(path):
    """元シートから各週・曜日・時間帯の出勤講師を読み取る（全講師、絞り込み前）"""
    wb = openpyxl.load_workbook(path)
    weeks = []
    
    # シート名でフィルタリング（「ブース表」を含むシートのみ対象）
    target_sheets = []
    for sn in wb.sheetnames:
        # 非表示シートはスキップ
        if wb[sn].sheet_state != 'visible':
            continue
        # 「ブース表」が含まれるシートのみ対象
        if 'ブース表' in sn:
            target_sheets.append(sn)

    # フォールバック: 「ブース表」を含むシートがなければメタシート以外の全可視シートを対象
    if not target_sheets:
        for sn in wb.sheetnames:
            if wb[sn].sheet_state != 'visible':
                continue
            if any(k in sn for k in META_KEYWORDS):
                continue
            target_sheets.append(sn)

    # Pass 1: 全講師フルネームを収集して同姓検出
    all_full_names = []
    for sn in target_sheets:
        ws = wb[sn]
        for day in DAYS:
            col = SRC_DAY_COLS[day]
            for start, tl, nb in SRC_TIME_SLOTS:
                for b in range(nb):
                    v = ws.cell(start+b*2, col).value
                    if v and str(v).strip():
                        all_full_names.append(str(v).strip())
    _build_name_map(all_full_names)

    # Pass 2: 通常パース
    for sn in target_sheets:
        ws = wb[sn]
        week = {}
        for day in DAYS:
            col = SRC_DAY_COLS[day]
            dt = {}
            for start, tl, nb in SRC_TIME_SLOTS:
                ts = TIME_SHORT[tl]
                teachers = []
                for b in range(nb):
                    t = to_short(ws.cell(start+b*2, col).value)
                    if t:
                        teachers.append(t)
                dt[ts] = teachers
            week[day] = dt
        
        # 週全体で講師が一人もいない場合はスキップ（空のシートを除外）
        has_teachers = False
        for d_data in week.values():
            for t_list in d_data.values():
                if t_list:
                    has_teachers = True
                    break
            if has_teachers: break
        
        if has_teachers:
            weeks.append(week)
    wb.close()
    return weeks

# ========== 元シート集約（講師回答ファイル → 週別出勤データ） ==========
SURVEY_TIME_ROWS = {10: '14:55', 11: '16:00', 12: '17:05', 13: '18:10', 14: '19:15', 15: '20:20'}
WEEKDAY_NORMALIZE = {
    '月曜日':'月','月曜':'月','月':'月','Mon':'月',
    '火曜日':'火','火曜':'火','火':'火','Tue':'火',
    '水曜日':'水','水曜':'水','水':'水','Wed':'水',
    '木曜日':'木','木曜':'木','木':'木','Thu':'木',
    '金曜日':'金','金曜':'金','金':'金','Fri':'金',
    '土曜日':'土','土曜':'土','土':'土','Sat':'土',
    '日曜日':'日','日曜':'日','日':'日','Sun':'日',
}

def _compute_month_week_map(year, month):
    """月の各日を週インデックス（1始まり）にマッピング。月〜土を1週とする。"""
    last_day = calendar.monthrange(year, month)[1]
    week_map = {}
    current_week = 0
    first_monday_seen = False
    for d in range(1, last_day + 1):
        dt = _dt.date(year, month, d)
        dow = dt.weekday()
        if dow == 6:  # 日曜はスキップ
            continue
        if dow == 0:  # 月曜で新しい週を開始
            current_week += 1
            first_monday_seen = True
        elif not first_monday_seen and current_week == 0:
            current_week = 1  # 月初が月曜以外の場合も第1週
        week_map[d] = current_week
    return week_map

def _get_merged_cell_value(ws, row, col):
    """結合セルの場合、左上セルの値を返す"""
    val = ws.cell(row, col).value
    if val is not None:
        return val
    from openpyxl.utils import get_column_letter
    coord = f'{get_column_letter(col)}{row}'
    for merge_range in ws.merged_cells.ranges:
        if coord in merge_range:
            return ws.cell(merge_range.min_row, merge_range.min_col).value
    return None

def _excel_serial_to_date(serial):
    """Excel シリアル値を date に変換"""
    try:
        return (_dt.datetime(1899, 12, 30) + _dt.timedelta(days=int(serial))).date()
    except Exception:
        return None

def parse_survey_file(file_path):
    """講師回答xlsxファイルを解析して講師名と出勤可能日時を返す"""
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # データシートを探す（「シート」を含むシート名、なければ先頭シート）
    data_sheet = None
    for sn in wb.sheetnames:
        if 'シート' in sn:
            data_sheet = sn
            break
    if not data_sheet:
        data_sheet = wb.sheetnames[0]

    ws = wb[data_sheet]

    # 講師名（row2, col2） — 結合セルにも対応
    raw_name = _get_merged_cell_value(ws, 2, 2)
    # 取得できなければ近傍セル (row1-3, col1-3) も探索
    if not raw_name:
        for r in range(1, 4):
            for c in range(1, 4):
                v = _get_merged_cell_value(ws, r, c)
                if v and isinstance(v, str) and len(v.strip()) >= 2:
                    # 数字だけ・記号だけのセルは除外
                    stripped = v.strip()
                    if not stripped.replace(' ', '').replace('\u3000', '').isdigit():
                        raw_name = v
                        break
            if raw_name:
                break
    # ファイル名からもフルネームを抽出（例: survey_井上 玲也クリスチャン_202603シート.xlsx）
    basename = os.path.basename(file_path)
    if basename.startswith('survey_'):
        basename = basename[7:]
    filename_name = basename.split('_')[0].strip()

    # セルの名前が単一語（名前のみ）でファイル名に姓名がある場合、ファイル名を優先
    if raw_name and filename_name:
        cell_parts = str(raw_name).replace('\u3000', ' ').split()
        fn_parts = filename_name.replace('\u3000', ' ').split()
        if len(cell_parts) == 1 and len(fn_parts) >= 2:
            raw_name = filename_name
    elif not raw_name and filename_name:
        raw_name = filename_name

    teacher_name = to_short(raw_name) if raw_name else None
    if not teacher_name:
        print(f"[survey] 講師名を検出できません: {file_path}", flush=True)
        return None

    # 日付から年月を取得し、週マップを構築
    year, month = None, None
    for c in range(3, ws.max_column + 1):
        v = _get_merged_cell_value(ws, 6, c)
        if isinstance(v, (_dt.datetime, _dt.date)):
            dt = v if isinstance(v, _dt.date) else v.date()
            year, month = dt.year, dt.month
            break
        elif isinstance(v, (int, float)) and v > 31:
            # Excel シリアル値の可能性
            dt = _excel_serial_to_date(v)
            if dt:
                year, month = dt.year, dt.month
                break

    # ヘッダー行から年月を推定（日付セルから取れなかった場合）
    if year is None or month is None:
        for r in range(1, 6):
            for c in range(1, ws.max_column + 1):
                v = _get_merged_cell_value(ws, r, c)
                if v and isinstance(v, str):
                    m = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月', v)
                    if m:
                        year, month = int(m.group(1)), int(m.group(2))
                        break
                    m2 = re.search(r'(\d{1,2})\s*月', v)
                    if m2 and year is None:
                        month = int(m2.group(1))
                        year = _dt.date.today().year
                        break
                elif isinstance(v, (_dt.datetime, _dt.date)):
                    dt = v if isinstance(v, _dt.date) else v.date()
                    year, month = dt.year, dt.month
                    break
            if year and month:
                break

    # ファイル名から年月を抽出するフォールバック（例: "202604シート" → 2026年4月）
    if year is None or month is None:
        fn_match = re.search(r'(\d{4})(\d{2})', os.path.basename(file_path))
        if fn_match:
            y, m = int(fn_match.group(1)), int(fn_match.group(2))
            if 2020 <= y <= 2100 and 1 <= m <= 12:
                year, month = y, m

    week_map = _compute_month_week_map(year, month) if year and month else {}

    # 列ヘッダー（日付・曜日・祝日フラグ）を読み取る
    columns = []
    j = 3
    consecutive_empty = 0
    while consecutive_empty < 3:
        # row 6: 日付, row 7: 曜日, row 9: 祝休日フラグ
        date_val = _get_merged_cell_value(ws, 6, j)
        if date_val is None or str(date_val).strip() == '':
            consecutive_empty += 1
            j += 1
            continue
        consecutive_empty = 0

        weekday_raw = str(_get_merged_cell_value(ws, 7, j) or '').strip()
        weekday = WEEKDAY_NORMALIZE.get(weekday_raw, weekday_raw)

        # 曜日が取れなかった場合はdateオブジェクトから推測
        resolved_date = None
        if isinstance(date_val, (_dt.datetime, _dt.date)):
            resolved_date = date_val if isinstance(date_val, _dt.date) else date_val.date()
        elif isinstance(date_val, (int, float)) and date_val > 31:
            resolved_date = _excel_serial_to_date(date_val)

        if weekday not in DAYS and resolved_date:
            wd_names = ['月','火','水','木','金','土','日']
            weekday = wd_names[resolved_date.weekday()]

        # 日付から週番号を算出（row 8 は曜日出現回数なので使わない）
        day_of_month = None
        if resolved_date:
            day_of_month = resolved_date.day
        elif isinstance(date_val, (int, float)) and 1 <= date_val <= 31:
            day_of_month = int(date_val)
        elif isinstance(date_val, str) and date_val.strip().isdigit():
            d = int(date_val.strip())
            if 1 <= d <= 31:
                day_of_month = d
        week_num = week_map.get(day_of_month)

        holiday = _get_merged_cell_value(ws, 9, j)

        columns.append({
            'col': j,
            'weekday': weekday,
            'week_num': week_num,
            'holiday': (holiday == 1 or str(holiday).strip() == '1'),
        })
        j += 1

    # 出勤可能時間帯を読み取る
    availability = []  # list of {weekday, week_num, time_str}
    for row, time_str in SURVEY_TIME_ROWS.items():
        for col_info in columns:
            if col_info['holiday']:
                continue
            if col_info['weekday'] == '日':
                continue
            val = ws.cell(row, col_info['col']).value
            if val == 1 or str(val).strip() == '1':
                availability.append({
                    'weekday': col_info['weekday'],
                    'week_num': col_info['week_num'],
                    'time': time_str,
                })

    wb.close()
    none_weeks = sum(1 for a in availability if a.get('week_num') is None)
    if none_weeks:
        print(f"[survey] WARNING: {teacher_name} — {none_weeks}/{len(availability)}コマで week_num=None (year={year}, month={month})", flush=True)
    print(f"[survey] parsed: {teacher_name} (full: {raw_name}) — {len(availability)}コマ, year={year}, month={month}, cols={len(columns)}", flush=True)

    return {
        'name': teacher_name,
        'full_name': str(raw_name).strip(),
        'availability': availability,
    }

def aggregate_surveys_to_weekly(survey_results):
    """複数の講師回答データを集約して週別出勤講師データを生成する
    Returns: load_weekly_teachers と同じ形式 — weeks[wi][day][ts] = [teacher, ...]
    """
    # インデックス構築: (week_num, day, time) → set of teacher names
    index = defaultdict(set)
    max_week = 0
    for sr in survey_results:
        fn = sr['name']  # 短縮名を使用（スケジュールの講師名と一致させるため）
        for a in sr['availability']:
            wn = a.get('week_num')
            if wn and wn > max_week:
                max_week = wn
            key = (wn, a['weekday'], a['time'])
            index[key].add(fn)
    if max_week == 0:
        max_week = 4  # fallback

    # week_num=None のデータを全週に配分（データ消失を防止）
    for key in list(index.keys()):
        if key[0] is None:
            _, day, time_str = key
            teachers = index[key]
            for wn in range(1, max_week + 1):
                index[(wn, day, time_str)].update(teachers)
            del index[key]

    weeks = []
    for wi in range(max_week):
        week = {}
        for day in DAYS:
            dt = {}
            for time_str in ALL_TIMES:
                ts = TIME_SHORT[time_str]
                dt[ts] = list(index.get((wi + 1, day, time_str), []))
            week[day] = dt
        weeks.append(week)

    return weeks

def _merge_weekly_teachers(base_wt, overlay_wt):
    """2つのweeklyTeachers構造をunionマージする。Returns: マージ済みの新しいリスト"""
    if not base_wt:
        return list(overlay_wt) if overlay_wt else []
    if not overlay_wt:
        return list(base_wt)
    max_weeks = max(len(base_wt), len(overlay_wt))
    merged = []
    for wi in range(max_weeks):
        bw = base_wt[wi] if wi < len(base_wt) else {}
        ow = overlay_wt[wi] if wi < len(overlay_wt) else {}
        week = {}
        for day in set(list(bw.keys()) + list(ow.keys())):
            bd = bw.get(day, {})
            od = ow.get(day, {})
            day_data = {}
            for ts in set(list(bd.keys()) + list(od.keys())):
                day_data[ts] = sorted(set(bd.get(ts, [])) | set(od.get(ts, [])))
            week[day] = day_data
        merged.append(week)
    return merged

def _process_survey_files(file_list, session_dir):
    """講師回答ファイルを処理し、survey_results・errors・survey_name_mapを返す共通ヘルパー。
    Returns: (survey_results, errors, survey_name_map)
    """
    survey_results = []
    errors = []
    for f in file_list:
        fname = os.path.basename(f.filename or '')
        if not fname or '出力' in fname or 'メタデータ' in fname:
            continue
        ok, err = validate_file(f)
        if not ok:
            errors.append(f'{f.filename}: {err}')
            continue
        path = os.path.join(session_dir, 'survey_' + fname)
        try:
            f.save(path)
        except Exception as e:
            errors.append(f'{f.filename}: 保存失敗 - {e}')
            continue
        try:
            result = parse_survey_file(path)
            if result:
                survey_results.append(result)
                print(f"[survey] {f.filename} -> {result['name']} ({len(result['availability'])}コマ)", flush=True)
            else:
                errors.append(f'{f.filename}: 講師情報を読み取れません')
        except Exception as e:
            errors.append(f'{f.filename}: {str(e)}')
            traceback.print_exc()
    # 同姓講師を自動検出し、短縮名を再計算
    if survey_results:
        all_full = [sr['full_name'] for sr in survey_results if sr.get('full_name')]
        _build_name_map(all_full)
        for sr in survey_results:
            sr['name'] = to_short(sr.get('full_name', sr['name']))
    survey_name_map = {sr['name']: sr.get('full_name', '') for sr in survey_results}
    return survey_results, errors, survey_name_map

def generate_src_excel(weekly_teachers, output_path):
    """週別出勤講師データからブース表元シートExcel（1ブック・週別シート）を生成"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for wi, week_data in enumerate(weekly_teachers):
        ws = wb.create_sheet(f'第{wi+1}週')

        for start_row, time_str, num_booths in SRC_TIME_SLOTS:
            ts = TIME_SHORT[time_str]
            for day, col in SRC_DAY_COLS.items():
                teachers = week_data.get(day, {}).get(ts, [])
                for bi in range(min(len(teachers), num_booths)):
                    ws.cell(start_row + bi * 2, col, teachers[bi])

    wb.save(output_path)
    wb.close()

def select_teachers_for_day(day, day_data, booth_pref, wish_teachers_set, office_teacher=None):
    """
    1日分の全時間帯データから、ブース⑥まで（最大6名）に講師を絞り込む。
    - 教室業務担当(office_teacher)はブースから除外
    - 早い時間帯から出勤できる講師を優先
    - 希望講師(wish_teachers_set)はブース枠外でも配置可能（例外）
    - ブース希望がある講師はその番号に配置
    """
    times = SATURDAY_TIMES if day == '土' else WEEKDAY_TIMES
    ts_list = [TIME_SHORT[tl] for tl in times]

    # 各講師の最早出勤時間帯を計算（教室業務担当を除外）
    teacher_earliest = {}
    for ts in ts_list:
        for t in day_data.get(ts, []):
            if t == office_teacher:
                continue  # 教室業務担当はブースに入れない
            if t not in teacher_earliest:
                teacher_earliest[t] = ts

    if len(teacher_earliest) <= MAX_BOOTHS:
        selected = set(teacher_earliest.keys())
    else:
        ts_order = {'14':0, '16':1, '17':2, '18':3, '19':4, '20':5}
        ranked = sorted(teacher_earliest.items(), key=lambda x: ts_order.get(x[1], 99))
        selected = set(t for t, _ in ranked[:MAX_BOOTHS])

    # 希望講師は必ず含める（例外）
    for t in teacher_earliest:
        if t in wish_teachers_set:
            selected.add(t)

    # ブース希望を正確に反映した配置を生成
    # 1) 希望ブースがある講師を先にその番号に配置
    # 2) 残りの講師を空きブースに詰める
    def assign_booth_order(available_teachers):
        """available_teachersリストをブース希望に基づいて最大6スロットに配置"""
        slots = [None] * MAX_BOOTHS  # ブース①〜⑥

        # 希望ブースがある講師を先に配置
        remaining = list(available_teachers)
        for t in list(remaining):
            if t in booth_pref:
                bi = booth_pref[t] - 1  # 0-indexed
                if 0 <= bi < MAX_BOOTHS and slots[bi] is None:
                    slots[bi] = t
                    remaining.remove(t)

        # 残りを空きスロットに順番に詰める
        for t in remaining:
            for i in range(MAX_BOOTHS):
                if slots[i] is None:
                    slots[i] = t
                    break

        return [t for t in slots if t is not None]

    # 各講師の出勤範囲（最初〜最後の出勤コマ）を計算
    # 元シートの読み込み制限(nb=6)で中間コマが欠落する場合を補間する
    ts_order = {'14':0, '16':1, '17':2, '18':3, '19':4, '20':5}
    teacher_range = {}  # {teacher: (first_ord, last_ord)}
    for ts in ts_list:
        for t in day_data.get(ts, []):
            if t not in selected or t == office_teacher:
                continue
            o = ts_order.get(ts, 99)
            if t not in teacher_range:
                teacher_range[t] = (o, o)
            else:
                teacher_range[t] = (teacher_range[t][0], max(teacher_range[t][1], o))

    # 1日分のブース配置を1回だけ決定し、全時間帯で同じブース番号を維持する
    # （途中で別講師がそのブースに入らないようにする）
    all_day_teachers = [t for t in selected if t != office_teacher]
    day_booth_order = assign_booth_order(all_day_teachers)

    result = {}
    for ts in ts_list:
        cur_ord = ts_order.get(ts, 99)
        # day_dataに直接含まれる講師 OR 出勤範囲内（first〜last）の講師
        available = set()
        for t in day_booth_order:
            if t in teacher_range:
                first, last = teacher_range[t]
                if first <= cur_ord <= last:
                    available.add(t)
        # 固定ブース位置に基づいてリスト生成（出勤していないコマは空文字）
        booths = []
        for i, t in enumerate(day_booth_order):
            if t in available:
                booths.append(t)
            else:
                booths.append('')  # そのコマは不在だがブース位置を確保
        result[ts] = booths
    return result

def resolve_office_teacher(day, candidates, day_data):
    """教室業務担当を優先順位リストから決定する。
    - day_data（その週・その曜日の出勤講師データ）で出勤確認
    - 誰も出勤していなければ None（教室業務なし）
    """
    if isinstance(candidates, str):
        candidates = [candidates]
    for candidate in candidates:
        # day_data: {ts: [teacher, ...]} — いずれかの時間帯に出勤していれば可
        for ts, teachers in day_data.items():
            if candidate in teachers:
                return candidate
    return None

def load_holidays(booth_wb, num_weeks):
    """ブース表の教室業務行(row 5)から休塾日を検出する。
    Returns: [{day: True, ...}, ...] 各週の休塾日マップ
    """
    # 隠しシートとメタデータシートを除外
    week_sheets = []
    for sn in booth_wb.sheetnames:
        if any(k in sn for k in META_KEYWORDS): continue
        if booth_wb[sn].sheet_state != 'visible': continue
        week_sheets.append(sn)

    holidays = []
    for wi in range(min(num_weeks, len(week_sheets))):
        ws = booth_wb[week_sheets[wi]]
        h = {}
        for day, cols in DAY_COLS.items():
            val = ws.cell(5, cols[0]).value
            if val and '休塾' in str(val):
                h[day] = True
        holidays.append(h)
    # 足りない週は空辞書で埋める
    while len(holidays) < num_weeks:
        holidays.append({})
    return holidays

def load_holidays_from_files(week_file_paths):
    """週ファイルリストから休塾日を検出する。
    Returns: [{day: True, ...}, ...] 各週の休塾日マップ
    """
    holidays = []
    for wp in week_file_paths:
        wb = openpyxl.load_workbook(wp, read_only=True)
        week_sheets = [sn for sn in wb.sheetnames
                       if 'ブース表' in sn and wb[sn].sheet_state == 'visible']
        if week_sheets:
            ws = wb[week_sheets[0]]
            h = {}
            for day, cols in DAY_COLS.items():
                val = ws.cell(5, cols[0]).value
                if val and '休塾' in str(val):
                    h[day] = True
            holidays.append(h)
        else:
            holidays.append({})
        wb.close()
    return holidays

def extract_week_dates_from_files(week_file_paths):
    """週ファイルリストからシート名の日付を算出する。
    Returns: {'year':int, 'month':int, 'weeks':[ {day_name: day_number, ...}, ... ]}
    """
    year, month = None, None
    for wp in week_file_paths:
        wb = openpyxl.load_workbook(wp, read_only=True)
        for sn in wb.sheetnames:
            m = re.search(r'(\d{4})[./](\d{1,2})[./](\d{1,2})', sn)
            if m:
                year, month = int(m.group(1)), int(m.group(2))
                break
        wb.close()
        if year:
            break
    if not year:
        return None

    day_names = ['月','火','水','木','金','土']
    week_map = _compute_month_week_map(year, month)

    by_week = {}
    for day_num, week_num in week_map.items():
        dt = _dt.date(year, month, day_num)
        wd = dt.weekday()
        if wd < 6:
            if week_num not in by_week:
                by_week[week_num] = {}
            by_week[week_num][day_names[wd]] = day_num

    num_weeks = len(week_file_paths)
    weeks = []
    for wi in range(num_weeks):
        weeks.append(by_week.get(wi + 1, {}))
    return {'year': year, 'month': month, 'weeks': weeks}

# ========== スケジューラー ==========
def build_schedule(students, weekly_teachers, skills, office_rule, booth_pref, holidays=None, weights=None, week_dates=None):
    if weights is None:
        weights = dict(DEFAULT_WEIGHTS)
    remaining = {s['name']: dict(s['needs']) for s in students}
    smap = {s['name']: s for s in students}
    schedule = []
    office_teachers = []
    num_weeks = len(weekly_teachers)

    # 各週の有効な曜日を算出（部分週対応: 月初・月末で存在しない曜日を除外）
    if week_dates and week_dates.get('weeks'):
        valid_days_per_week = []
        for wi in range(num_weeks):
            if wi < len(week_dates['weeks']):
                valid_days_per_week.append(set(week_dates['weeks'][wi].keys()))
            else:
                valid_days_per_week.append(set(DAYS))
    else:
        valid_days_per_week = [set(DAYS)] * num_weeks

    # 全生徒の希望講師を集約
    wish_teachers_set = set()
    for s in students:
        wish_teachers_set.update(s['wish_teachers'])

    for wi in range(num_weeks):
        ot = {}
        valid_days = valid_days_per_week[wi]
        for d in DAYS:
            # 存在しない曜日チェック（部分週: 月初・月末）
            if d not in valid_days:
                ot[d] = '休塾日'
            # 休塾日チェック
            elif holidays and wi < len(holidays) and holidays[wi].get(d):
                ot[d] = '休塾日'
            else:
                candidates = office_rule.get(d, [])
                d_data = weekly_teachers[wi].get(d, {})
                ot[d] = resolve_office_teacher(d, candidates, d_data)
        office_teachers.append(ot)
        ws = {}
        for day in DAYS:
            ds = {}
            times = SATURDAY_TIMES if day=='土' else WEEKDAY_TIMES
            ot_teacher = ot.get(day)
            if ot_teacher == '休塾日':
                # 休塾日は空ブースのみ（講師を配置しない）
                for tl in times:
                    ts = TIME_SHORT[tl]
                    ds[ts] = [{'teacher':'', 'slots':[]} for _ in range(MAX_BOOTHS)]
            else:
                # 講師選抜（ブース⑥まで、早い時間帯優先、教室業務担当除外）
                day_data = weekly_teachers[wi].get(day, {})
                filtered = select_teachers_for_day(day, day_data, booth_pref, wish_teachers_set, ot_teacher)
                for tl in times:
                    ts = TIME_SHORT[tl]
                    tlist = filtered.get(ts, [])
                    booths = [{'teacher':t, 'slots':[]} for t in tlist]
                    # 常にMAX_BOOTHS(6)ブース分のデータを確保
                    while len(booths) < MAX_BOOTHS:
                        booths.append({'teacher':'', 'slots':[]})
                    ds[ts] = booths
            ws[day] = ds
        schedule.append(ws)

    # ---- 配置インデックス: O(1)ルックアップ用 ----
    # idx_placed_days[wi][name][subj] = set of days
    # idx_student_slots[wi][name] = set of (day, ts)
    # idx_any_days[wi][name] = set of days
    idx_placed_days = [{} for _ in range(num_weeks)]
    idx_student_slots = [{} for _ in range(num_weeks)]
    idx_any_days = [{} for _ in range(num_weeks)]

    def _update_index(wi, name, subj, day, ts):
        """配置成功時にインデックスを更新"""
        idx_placed_days[wi].setdefault(name, {}).setdefault(subj, set()).add(day)
        idx_student_slots[wi].setdefault(name, set()).add((day, ts))
        idx_any_days[wi].setdefault(name, set()).add(day)

    def get_placed_days(ws_unused, name, subj, wi):
        return idx_placed_days[wi].get(name, {}).get(subj, set())

    def get_student_slots(ws_unused, name, wi):
        return idx_student_slots[wi].get(name, set())

    def get_any_placed_days(ws_unused, name, wi):
        return idx_any_days[wi].get(name, set())

    def get_teacher_booth(ws, day, teacher):
        for ts, booths in ws.get(day,{}).items():
            for bi,b in enumerate(booths):
                if b['teacher']==teacher and b['slots']:
                    return bi
        return None

    def check_booth(booth, bi, s, day, subj, ws):
        t = booth['teacher']
        if not t or len(booth['slots'])>=2: return False
        if t in s['ng_teachers']: return False
        if not can_teach(t, s['grade'], subj, skills): return False
        # 同一ブース内のNG生徒チェック
        for g2,sn2,sb2 in booth['slots']:
            if sn2 in s['ng_students']: return False
            other = smap.get(sn2)
            if other and s['name'] in other.get('ng_students',[]): return False
        # 隣接ブースチェックは廃止（同一ブースのみNGとする要望により）
        
        eb = get_teacher_booth(ws, day, t)
        if eb is not None and eb != bi: return False
        return True

    def place_student(ws, s, day, ts, subj):
        if day not in ws or ts not in ws[day]: return False
        booths = list(enumerate(ws[day][ts]))
        wish = s.get('wish_teachers', [])
        if wish:
            # 希望講師のブースを先に試す
            wish_first = sorted(booths, key=lambda x: (0 if x[1]['teacher'] in wish else 1))
            booths = wish_first
        for bi,b in booths:
            if check_booth(b, bi, s, day, subj, ws):
                b['slots'].append((s['grade'],s['name'],subj))
                return True
        return False

    def find_slot(ws, s, subj, placed_days, existing, wi, any_placed_days):
        cands = []
        checked_avail = False
        reject_full = 0
        reject_ng = 0
        reject_skill = 0
        reject_other = 0
        for day in DAYS:
            if day not in valid_days_per_week[wi]: continue  # 存在しない曜日をスキップ
            if day in placed_days: continue  # 同一科目の同曜日配置を防止
            # NG日程チェック: 配置自体は許可するがペナルティ
            is_ng_date = (wi, day) in s.get('ng_dates', set())
            
            times = SATURDAY_TIMES if day=='土' else WEEKDAY_TIMES
            for tl in times:
                ts = TIME_SHORT[tl]
                is_primary = s['avail'] is None or (day,ts) in s['avail']
                is_backup = (not is_primary) and s.get('backup_avail') and (day,ts) in s['backup_avail']
                if not is_primary and not is_backup: continue
                checked_avail = True
                if (day,ts) in existing: continue
                if ts not in ws.get(day,{}): continue
                for bi,b in enumerate(ws[day][ts]):
                    t = b['teacher']
                    if not t: continue
                    if len(b['slots'])>=2:
                        reject_full += 1
                        continue
                    if t in s['ng_teachers']:
                        reject_ng += 1
                        continue
                    if not can_teach(t, s['grade'], subj, skills):
                        reject_skill += 1
                        continue
                    if not check_booth(b, bi, s, day, subj, ws):
                        reject_other += 1
                        continue
                    sc = 0
                    # NG日程は大きくペナルティ（配置は可能）
                    if is_ng_date: sc += weights['ng_date']
                    # 予備時間はペナルティ（希望時間を優先）
                    if is_backup: sc += weights['backup_time']
                    # 同曜日に既に別科目が配置されている場合
                    # 連続コマを強く推奨、飛び石は回避
                    existing_on_day = [t_ for d_, t_ in existing if d_ == day]
                    if existing_on_day:
                        # 現在の時刻のインデックスを取得
                        try:
                            # timesは '16:00' 等の形式リスト
                            # tl は現在ループ中の時刻文字列 ('16:00')
                            curr_idx = times.index(tl)

                            is_continuous = False
                            for et_short in existing_on_day:
                                et_long = TIME_SHORT_REV.get(et_short)
                                if et_long in times:
                                    ex_idx = times.index(et_long)
                                    diff = abs(curr_idx - ex_idx)
                                    if diff == 1:
                                        sc += weights['continuous_block']
                                        is_continuous = True
                                    elif diff > 1:
                                        sc += weights['skip_interval']
                        except ValueError:
                            pass

                    if day in any_placed_days:
                        day_count = len(existing_on_day)
                        if day_count < 2:
                            sc += weights['same_day_2nd']
                        else:
                            sc += weights['same_day_3plus']
                    if b['teacher'] in s['wish_teachers']: sc += weights['wish_teacher']
                    if t in booth_pref and booth_pref[t]==bi+1: sc += weights['booth_pref']
                    if len(b['slots'])==0: sc += weights['empty_booth']
                    cands.append((sc, day, ts, bi))
        if not cands:
            if not checked_avail:
                reason = '希望時間帯なし'
            elif reject_skill:
                reason = '指導可能な講師不在'
            elif reject_ng:
                reason = 'NG講師'
            elif reject_other:
                reason = 'NG生徒/ブース制約'
            elif reject_full:
                reason = 'ブース満席'
            else:
                reason = '空きコマなし'
            return None, reason
        cands.sort(key=lambda x:-x[0])
        best_sc = cands[0][0]
        bests = [c for c in cands if c[0]==best_sc]
        ch = random.choice(bests)
        return (ch[1], ch[2], ch[3]), None

    def distribute(total, weeks):
        if weeks <= 0:
            return []
        t = [total//weeks]*weeks
        for i in range(total%weeks): t[i] += 1
        random.shuffle(t)
        return t

    # Phase1: 固定授業（必要コマ数を超えても配置する — 固定曜日は全有効週に配置）
    for s in students:
        for day, ts_str, subj in s['fixed']:
            for wi in range(num_weeks):
                if day not in valid_days_per_week[wi]: continue  # 存在しない曜日をスキップ
                if (wi, day) in s.get('ng_dates', set()): continue
                if place_student(schedule[wi], s, day, ts_str, subj):
                    if subj in remaining[s['name']]:
                        remaining[s['name']][subj] -= 1
                    _update_index(wi, s['name'], subj, day, ts_str)

    # viable_slots: 週wiにおいて生徒sの希望時間帯に担当可能講師がいるスロット数
    # avail=None(制限なし)は999を返す。制約が多い生徒ほど小さい値になる。
    def viable_slots(s, wi):
        if s['avail'] is None:
            return 999
        wt = weekly_teachers[wi] if wi < len(weekly_teachers) else {}
        count = 0
        for day, ts in s['avail']:
            if day not in valid_days_per_week[wi]: continue  # 存在しない曜日をスキップ
            for b in wt.get(day, {}).get(ts, []):
                t = b.get('teacher', '') if isinstance(b, dict) else b
                if t and t not in s.get('ng_teachers', set()):
                    if any(can_teach(t, s['grade'], subj, skills) for subj in s['needs']):
                        count += 1
                        break
        return count

    # Phase2: 通常配置（希望講師ありの生徒を先に全て配置してから、その他の生徒を配置）
    # 週単位でviable_slots数（担当可能講師のある希望スロット数）が少ない順に配置
    all_students = sorted(students, key=lambda s: sum(s['needs'].values()))
    wish_order = [s for s in all_students if s['wish_teachers']]
    no_wish_order = [s for s in all_students if not s['wish_teachers']]
    unplaced_reasons = {}  # (name, subj) -> reason

    def _place_phase2(student_list):
        # 週ごとの配置数を事前に決定（distribute）
        dist = {}
        for s in student_list:
            dist[s['name']] = {}
            for subj in s['needs']:
                still = remaining[s['name']].get(subj, 0)
                dist[s['name']][subj] = distribute(still, num_weeks) if still > 0 else [0] * num_weeks

        for wi in range(num_weeks):
            # この週において担当可能スロット数が少ない順（最制約優先）にソート
            sorted_list = sorted(student_list, key=lambda s: viable_slots(s, wi))
            for s in sorted_list:
                for subj in s['needs']:
                    n = dist[s['name']][subj][wi]
                    for _ in range(n):
                        if remaining[s['name']].get(subj, 0) <= 0:
                            break
                        pd = get_placed_days(None, s['name'], subj, wi)
                        ex = get_student_slots(None, s['name'], wi)
                        apd = get_any_placed_days(None, s['name'], wi)
                        best, reason = find_slot(schedule[wi], s, subj, pd, ex, wi, apd)
                        if best:
                            day, ts, bi = best
                            schedule[wi][day][ts][bi]['slots'].append((s['grade'], s['name'], subj))
                            remaining[s['name']][subj] -= 1
                            _update_index(wi, s['name'], subj, day, ts)
                        elif reason:
                            unplaced_reasons[(s['name'], subj)] = reason

    # Phase2a: 希望講師ありの生徒を先に全て配置
    _place_phase2(wish_order)
    # Phase2b: 希望講師なしの生徒を配置
    _place_phase2(no_wish_order)

    # Phase3: 未配置リトライ（distribute で割り当てられなかった週にも配置を試行）
    for s in wish_order + no_wish_order:
        for subj in s['needs']:
            still = remaining[s['name']].get(subj, 0)
            if still <= 0: continue
            for wi in range(num_weeks):
                if remaining[s['name']].get(subj, 0) <= 0: break
                pd = get_placed_days(None, s['name'], subj, wi)
                ex = get_student_slots(None, s['name'], wi)
                apd = get_any_placed_days(None, s['name'], wi)
                best, reason = find_slot(schedule[wi], s, subj, pd, ex, wi, apd)
                if best:
                    day, ts, bi = best
                    schedule[wi][day][ts][bi]['slots'].append((s['grade'],s['name'],subj))
                    remaining[s['name']][subj] -= 1
                    _update_index(wi, s['name'], subj, day, ts)
                elif reason:
                    unplaced_reasons[(s['name'], subj)] = reason

    # Phase4: スワップ最適化（希望時間帯遵守率向上）
    # backup スロットにいる生徒を primary スロットの生徒とスワップして遵守率を改善。
    # 最大 MAX_SWAP_ITER 回繰り返し、スワップがゼロになった時点で早期終了。
    MAX_SWAP_ITER = 10

    def _is_primary_slot(s, day, ts):
        return s['avail'] is None or (day, ts) in s['avail']

    def _remove_index_entry(wi, name, subj, day, ts):
        """配置インデックスから1エントリを削除する"""
        idx_placed_days[wi].setdefault(name, {}).setdefault(subj, set()).discard(day)
        idx_student_slots[wi].setdefault(name, set()).discard((day, ts))
        if not any(d == day for d, _ in idx_student_slots[wi].get(name, set())):
            idx_any_days[wi].setdefault(name, set()).discard(day)

    for _iter in range(MAX_SWAP_ITER):
        total_swaps = 0
        for wi in range(num_weeks):
            ws = schedule[wi]
            # 配置済みエントリを収集: (day, ts, bi, slot_idx, student, subj)
            placed = []
            for day in DAYS:
                for ts, booths in ws.get(day, {}).items():
                    for bi, b in enumerate(booths):
                        for si, (grade, name, subj) in enumerate(b['slots']):
                            s = smap.get(name)
                            if s:
                                placed.append((day, ts, bi, si, s, subj))

            swapped = set()  # このイテレーションで処理済みのエントリインデックス
            for i in range(len(placed)):
                if i in swapped:
                    continue
                day_a, ts_a, bi_a, si_a, s_a, subj_a = placed[i]
                # 固定授業は交換しない
                if any((day_a, ts_a) == (fd, ft) for fd, ft, _ in s_a.get('fixed', [])):
                    continue
                # すでに primary なら交換不要
                if _is_primary_slot(s_a, day_a, ts_a):
                    continue

                for j in range(len(placed)):
                    if j == i or j in swapped:
                        continue
                    day_b, ts_b, bi_b, si_b, s_b, subj_b = placed[j]
                    if s_a['name'] == s_b['name']:
                        continue
                    # 同一ブース内のエントリは交換しない（同一オブジェクトへの二重pop防止）
                    if day_a == day_b and ts_a == ts_b and bi_a == bi_b:
                        continue
                    # 固定授業は交換しない
                    if any((day_b, ts_b) == (fd, ft) for fd, ft, _ in s_b.get('fixed', [])):
                        continue

                    prim_b = _is_primary_slot(s_b, day_b, ts_b)
                    new_prim_a = _is_primary_slot(s_a, day_b, ts_b)
                    new_prim_b = _is_primary_slot(s_b, day_a, ts_a)
                    # prim_a は False（上でチェック済み）
                    if int(new_prim_a) + int(new_prim_b) <= int(prim_b):
                        continue  # 合計 primary 数が増えない

                    # 同曜日・同科目重複チェック
                    if day_b != day_a:
                        if day_b in idx_placed_days[wi].get(s_a['name'], {}).get(subj_a, set()):
                            continue
                        if day_a in idx_placed_days[wi].get(s_b['name'], {}).get(subj_b, set()):
                            continue

                    # 同一 (day, ts) への重複配置チェック
                    if (day_b, ts_b) in idx_student_slots[wi].get(s_a['name'], set()):
                        continue
                    if (day_a, ts_a) in idx_student_slots[wi].get(s_b['name'], set()):
                        continue

                    b_a = ws[day_a][ts_a][bi_a]
                    b_b = ws[day_b][ts_b][bi_b]

                    # 一時的に両エントリを取り外してブース制約を確認
                    slot_a = b_a['slots'].pop(si_a)
                    slot_b = b_b['slots'].pop(si_b)

                    ok_a = check_booth(b_b, bi_b, s_a, day_b, subj_a, ws)
                    ok_b = check_booth(b_a, bi_a, s_b, day_a, subj_b, ws)

                    if ok_a and ok_b:
                        # スワップ実行
                        b_b['slots'].insert(si_b, slot_a)
                        b_a['slots'].insert(si_a, slot_b)
                        _remove_index_entry(wi, s_a['name'], subj_a, day_a, ts_a)
                        _remove_index_entry(wi, s_b['name'], subj_b, day_b, ts_b)
                        _update_index(wi, s_a['name'], subj_a, day_b, ts_b)
                        _update_index(wi, s_b['name'], subj_b, day_a, ts_a)
                        swapped.add(i)
                        swapped.add(j)
                        total_swaps += 1
                        break
                    else:
                        # 元に戻す
                        b_a['slots'].insert(si_a, slot_a)
                        b_b['slots'].insert(si_b, slot_b)

        if total_swaps == 0:
            break

    unplaced = []
    for s in students:
        for subj, cnt in remaining[s['name']].items():
            if cnt > 0:
                reason = unplaced_reasons.get((s['name'], subj), '')
                unplaced.append({'grade':s['grade'],'name':s['name'],'subject':subj,'count':cnt,'reason':reason})

    return schedule, unplaced, office_teachers

def extract_week_dates(booth_wb, num_weeks):
    """ブース表シート名から各週・各曜日の日付を算出する。
    _compute_month_week_map を使用して正確な週境界で日付をマッピングする。
    Returns: {'year':int, 'month':int, 'weeks':[ {day_name: day_number, ...}, ... ]}
    """
    week_sheets = [sn for sn in booth_wb.sheetnames if not any(k in sn for k in META_KEYWORDS)]

    year, month = None, None
    for sn in week_sheets:
        m = re.search(r'(\d{4})[./](\d{1,2})[./](\d{1,2})', sn)
        if m:
            year, month = int(m.group(1)), int(m.group(2))
            break
    if not year:
        return None

    day_names = ['月','火','水','木','金','土']
    week_map = _compute_month_week_map(year, month)

    # 週番号ごとに日付をグループ化
    by_week = {}
    for day_num, week_num in week_map.items():
        dt = _dt.date(year, month, day_num)
        wd = dt.weekday()  # 0=Mon ... 5=Sat
        if wd < 6:
            if week_num not in by_week:
                by_week[week_num] = {}
            by_week[week_num][day_names[wd]] = day_num

    # 0-indexed リストに変換 (wi=0 → week 1)
    weeks = []
    for wi in range(num_weeks):
        weeks.append(by_week.get(wi + 1, {}))
    return {'year': year, 'month': month, 'weeks': weeks}

# ========== Excel出力 ==========
def _copy_worksheet_fast(src_ws, dst_ws, on_row_done=None):
    """Cross-workbook worksheet copy with style caching.
    on_row_done: optional callback called after each row is copied.
    Uses openpyxl internal _style tuple as cache key (avoids slow str() conversion).
    """
    style_cache = {}
    for row in src_ws.iter_rows():
        for cell in row:
            dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
            dst_cell.value = cell.value
            if cell.has_style:
                style_key = cell._style
                if style_key not in style_cache:
                    style_cache[style_key] = (
                        copy(cell.font), copy(cell.border), copy(cell.fill),
                        cell.number_format, copy(cell.protection), copy(cell.alignment)
                    )
                cached = style_cache[style_key]
                dst_cell.font = cached[0]
                dst_cell.border = cached[1]
                dst_cell.fill = cached[2]
                dst_cell.number_format = cached[3]
                dst_cell.protection = cached[4]
                dst_cell.alignment = cached[5]
        if on_row_done:
            on_row_done()
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
    for row_num, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = dim.height

def _write_schedule_to_ws(ws, wsched, office_data, on_batch_done=None):
    """1つの週シートにスケジュールデータを書き込む共通処理
    on_batch_done: optional callback called after each time_label is written.
    """
    teacher_font = Font(name='MS PGothic', size=8)
    teacher_align = Alignment(textRotation=255, vertical='center', horizontal='center')
    data_font = Font(name='MS PGothic', size=11)
    data_align = Alignment(vertical='center', horizontal='center')

    # クリア（結合は解除しない）
    for tl, (sr, nb) in LAYOUT.items():
        for b in range(nb):
            r1, r2 = sr+b*2, sr+b*2+1
            for day in DAYS:
                _, lc, gc, sc, sjc = DAY_COLS[day]
                try: ws.cell(r1, lc).value = None
                except Exception: pass
                for c in [gc, sc, sjc]:
                    for r in [r1, r2]:
                        try: ws.cell(r, c).value = None
                        except Exception: pass

    # 書き込み
    for tl, (sr, nb) in LAYOUT.items():
        ts = TIME_SHORT[tl]
        for day in DAYS:
            _, lc, gc, sc, sjc = DAY_COLS[day]
            booths = wsched.get(day,{}).get(ts,[])
            for bi in range(min(nb, len(booths))):
                r1, r2 = sr+bi*2, sr+bi*2+1
                b = booths[bi]
                if b['teacher']:
                    cell = ws.cell(r1, lc)
                    cell.value = b['teacher']
                    cell.font = teacher_font
                    cell.alignment = teacher_align
                if len(b['slots'])>=1:
                    g,sn,subj = b['slots'][0]
                    for c, v in [(gc,g),(sc,sn),(sjc,subj)]:
                        cell = ws.cell(r1,c)
                        cell.value = v
                        cell.font = data_font
                        cell.alignment = data_align
                if len(b['slots'])>=2:
                    g2,sn2,subj2 = b['slots'][1]
                    for c, v in [(gc,g2),(sc,sn2),(sjc,subj2)]:
                        cell = ws.cell(r2,c)
                        cell.value = v
                        cell.font = data_font
                        cell.alignment = data_align
        if on_batch_done:
            on_batch_done()

    # 教室業務・チューター
    holiday_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    holiday_font = Font(name='MS PGothic', color='333333', bold=True, size=11)
    holiday_align = Alignment(horizontal='center', vertical='center')
    for day in DAYS:
        bc = DAY_COLS[day][0]
        t = office_data.get(day)
        if t:
            cell = ws.cell(5, bc, t)
            if t == '休塾日':
                cell.fill = holiday_fill
                cell.font = holiday_font
                cell.alignment = holiday_align
            for tr in TUTOR_ROWS:
                try:
                    c = ws.cell(tr, bc, t)
                    if t == '休塾日':
                        c.fill = holiday_fill
                        c.font = holiday_font
                        c.alignment = holiday_align
                except Exception: pass
            if t == '休塾日':
                all_cols = DAY_COLS[day]
                for tl, (sr, nb) in LAYOUT.items():
                    for b_i in range(nb):
                        r1, r2 = sr + b_i * 2, sr + b_i * 2 + 1
                        for col in all_cols:
                            for r in [r1, r2]:
                                try:
                                    cell = ws.cell(r, col)
                                    cell.fill = holiday_fill
                                except Exception: pass
    if on_batch_done:
        on_batch_done()

def write_excel(schedule, unplaced, office_teachers, booth_path, output_path, week_file_paths=None, progress_fn=None):
    num_weeks = len(schedule)

    # --- セル数ベース進捗 ---
    # フェーズ配分:  準備 0-10%  |  週処理 10-85%  |  仕上げ 85-100%
    PHASE_WEEKS_START = 10
    PHASE_WEEKS_END = 85
    PHASE_FINAL = 85
    _last_pct = [-1]

    def _emit(pct, msg):
        pct = max(0, int(pct))
        if progress_fn and pct != _last_pct[0]:
            _last_pct[0] = pct
            progress_fn(pct, msg)

    _emit(0, '準備中...')
    week_range = PHASE_WEEKS_END - PHASE_WEEKS_START  # 75

    # 1週あたりのスケジュール書き込みバッチ数 (time_labels + office = 7)
    SCHED_BATCHES = len(LAYOUT) + 1  # 6 time_labels + 1 office = 7

    if week_file_paths:
        # --- 1週目高速化: ファイルコピーでWBを作成し、セルコピーをスキップ ---
        _emit(PHASE_WEEKS_START, '第1週を読み込み中...')
        tmp_path = output_path.replace('.xlsx', '_tmp.xlsx')
        shutil.copy2(week_file_paths[0], tmp_path)
        wb = openpyxl.load_workbook(tmp_path)
        try:
            os.remove(tmp_path)
        except Exception:
            pass

        # ブース表シートを特定し、不要シートを削除
        target_sn = None
        for sn in wb.sheetnames:
            if 'ブース表' in sn and wb[sn].sheet_state == 'visible':
                target_sn = sn
                break
        for sn in list(wb.sheetnames):
            if sn != target_sn:
                del wb[sn]

        # 1週目: スケジュール書き込みのみ（セルコピー不要）
        w_base_0 = PHASE_WEEKS_START
        w_alloc_0 = week_range / num_weeks
        batches_done_0 = [0]

        def _on_batch_0(bd=batches_done_0, tb=SCHED_BATCHES, base=w_base_0, alloc=w_alloc_0):
            bd[0] += 1
            _emit(base + (bd[0] / tb) * alloc, '第1週 スケジュール書き込み中')

        if target_sn:
            ot0 = office_teachers[0] if office_teachers else {}
            _write_schedule_to_ws(wb[target_sn], schedule[0], ot0, on_batch_done=_on_batch_0)

        # 2週目以降: 従来通りセルコピー
        for wi in range(1, min(num_weeks, len(week_file_paths))):
            w_base = PHASE_WEEKS_START + (wi / num_weeks) * week_range
            w_alloc = week_range / num_weeks

            _emit(w_base, f'第{wi+1}週を読み込み中...')
            week_wb = openpyxl.load_workbook(week_file_paths[wi])
            src_sn = None
            for sn in week_wb.sheetnames:
                if 'ブース表' in sn and week_wb[sn].sheet_state == 'visible':
                    src_sn = sn
                    break
            if not src_sn:
                week_wb.close()
                continue

            src_ws = week_wb[src_sn]
            dst_ws = wb.create_sheet(src_sn)

            # セルコピー (週配分の80%) — 行数ベース進捗
            copy_alloc = w_alloc * 0.8
            total_rows = max(src_ws.max_row or 1, 1)
            rows_done = [0]

            def _on_row(rd=rows_done, tr=total_rows, base=w_base, ca=copy_alloc, w=wi):
                rd[0] += 1
                if rd[0] % 10 == 0 or rd[0] == tr:
                    _emit(base + (rd[0] / tr) * ca, f'第{w+1}週 セルコピー中 ({rd[0]}/{tr}行)')

            _copy_worksheet_fast(src_ws, dst_ws, on_row_done=_on_row)
            week_wb.close()

            # スケジュール書き込み (週配分の20%) — バッチベース進捗
            sched_base = w_base + copy_alloc
            sched_alloc = w_alloc * 0.2
            batches_done = [0]

            def _on_batch(bd=batches_done, tb=SCHED_BATCHES, sb=sched_base, sa=sched_alloc, w=wi):
                bd[0] += 1
                _emit(sb + (bd[0] / tb) * sa, f'第{w+1}週 スケジュール書き込み中')

            ot = office_teachers[wi] if wi < len(office_teachers) else {}
            _write_schedule_to_ws(dst_ws, schedule[wi], ot, on_batch_done=_on_batch)

        week_sheets = [sn for sn in wb.sheetnames]
    elif booth_path:
        _emit(PHASE_WEEKS_START, 'テンプレートを読み込み中...')
        wb = openpyxl.load_workbook(booth_path)
        # メタシート・不要シートを特定してワークブックから削除（save高速化）
        remove_sheets = []
        week_sheets = []
        for sn in wb.sheetnames:
            if any(k in sn for k in META_KEYWORDS):
                remove_sheets.append(sn)
            elif sn.startswith('_schedule_data') or sn == '未配置コマ':
                remove_sheets.append(sn)
            else:
                week_sheets.append(sn)
        for sn in remove_sheets:
            del wb[sn]
        num_weeks = min(num_weeks, len(week_sheets))

        for wi in range(num_weeks):
            w_base = PHASE_WEEKS_START + (wi / num_weeks) * week_range
            w_alloc = week_range / num_weeks
            batches_done = [0]

            def _on_batch(bd=batches_done, tb=SCHED_BATCHES, base=w_base, alloc=w_alloc, w=wi):
                bd[0] += 1
                _emit(base + (bd[0] / tb) * alloc, f'第{w+1}週 スケジュール書き込み中')

            ws = wb[week_sheets[wi]]
            ot = office_teachers[wi] if wi < len(office_teachers) else {}
            _write_schedule_to_ws(ws, schedule[wi], ot, on_batch_done=_on_batch)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    # --- 仕上げフェーズ (85-100%) ---
    # 残存する不要シートを削除（week_file_pathsモードの出力ファイル再生成時など）
    for old_sn in list(wb.sheetnames):
        if old_sn == '未配置コマ' or old_sn.startswith('_schedule_data'):
            del wb[old_sn]

    # 保存 (85-100%) — サブスレッド + キープアライブ
    _emit(85, 'ファイルを保存中...')
    save_error = [None]
    save_done = threading.Event()
    def _do_save():
        try:
            wb.save(output_path)
        except Exception as e:
            save_error[0] = e
        finally:
            save_done.set()
    save_thread = threading.Thread(target=_do_save, daemon=True)
    save_thread.start()
    pct = 92
    while not save_done.wait(timeout=3.0):
        pct = min(pct + 1, 99)
        _emit(pct, 'ファイルを保存中...')
    save_thread.join(timeout=5)
    if save_error[0]:
        raise RuntimeError(f'Excelファイルの保存に失敗しました: {save_error[0]}') from save_error[0]
    wb.close()
    if progress_fn:
        progress_fn(100, '完了')

# ========== Error Handlers ==========
@app.errorhandler(413)
def request_entity_too_large(error):
    print(f"[error] 413 Request Entity Too Large", flush=True)
    return jsonify({'error': 'ファイルサイズが上限(10MB)を超えています。ファイルを確認してください。'}), 413

# ========== API ==========
@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/api/teachers')
@login_required
def get_teachers():
    """アップロード済みブース表から講師名一覧とブース希望を返す"""
    sd = get_session_data()
    files = sd.get('files', {})
    if 'booth' not in files:
        return jsonify({'error': 'ブース表がアップロードされていません'}), 400
    try:
        wb = openpyxl.load_workbook(files['booth'], data_only=True)
        skills = load_teacher_skills(wb)
        booth_pref = load_booth_pref(wb)
        if not booth_pref:
            booth_pref = dict(DEFAULT_BOOTH_PREF)
        wb.close()
        return jsonify({
            'teachers': sorted(skills.keys()),
            'boothPref': booth_pref,
        })
    except Exception as e:
        app.logger.error(f'teachers error: {traceback.format_exc()}')
        return jsonify({'error': '内部エラーが発生しました'}), 500

@app.route('/api/upload', methods=['POST'])
@login_required
def upload():
    sd = get_session_data()
    print(f"[upload] sid={sd.get('_sid','?')}, existing_files={list(sd.get('files',{}).keys())}", flush=True)
    saved = {}
    for key in ['src','booth']:
        f = request.files.get(key)
        if f:
            ok, err = validate_file(f)
            if not ok:
                print(f"[upload] ERROR validation failed: key={key}, filename={f.filename}, error={err}", flush=True)
                return jsonify({'error': err}), 400
            path = os.path.join(sd['dir'], key + '_' + os.path.basename(f.filename))
            try:
                f.save(path)
            except Exception as e:
                tb = traceback.format_exc()
                print(f"[upload] ERROR saving file: key={key}, filename={f.filename}, path={path}, error={e}\n{tb}", flush=True)
                return jsonify({'error': f'ファイル保存に失敗しました ({key}: {f.filename}): {e}'}), 500
            if not os.path.exists(path):
                print(f"[upload] ERROR file not found after save: key={key}, path={path}", flush=True)
                return jsonify({'error': f'ファイル保存後にファイルが見つかりません ({key}: {f.filename})'}), 500
            size = os.path.getsize(path)
            saved[key] = path
            print(f"[upload] saved {key} -> {path} (size={size}bytes)", flush=True)
    if not saved:
        print(f"[upload] WARNING no files received in request", flush=True)
        return jsonify({'error': 'アップロードするファイルが含まれていません'}), 400
    sd['files'] = {**sd.get('files',{}), **saved}
    save_session_files(sd)
    return jsonify({'ok': True, 'files': {k: (os.path.basename(v) if isinstance(v, str) else [os.path.basename(p) for p in v]) for k, v in sd.get('files', {}).items()}})

@app.route('/api/upload_surveys', methods=['POST'])
@login_required
def upload_surveys():
    """講師回答ファイル（複数）をアップロード → 集約 → 元シートを自動生成"""
    try:
        return _upload_surveys_impl()
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': f'サーバーエラー: {str(e)}'}), 500

def _upload_surveys_impl():
    sd = get_session_data()
    files = request.files.getlist('surveys')
    print(f"[survey] received {len(files)} files: {[f.filename for f in files]}", flush=True)
    if not files or all(not f.filename for f in files):
        return jsonify({'error': '講師回答ファイルが含まれていません'}), 400

    survey_results, errors, survey_name_map = _process_survey_files(files, sd['dir'])

    if not survey_results:
        detail_msg = '有効な講師回答ファイルがありません'
        if errors:
            detail_msg += '（' + '; '.join(errors[:5]) + '）'
        return jsonify({'error': detail_msg, 'details': errors}), 400

    sd['survey_name_map'] = survey_name_map
    save_session_files(sd)

    # 集約して元シートExcelを生成
    weekly_teachers = aggregate_surveys_to_weekly(survey_results)
    src_path = os.path.join(sd['dir'], 'generated_src.xlsx')
    generate_src_excel(weekly_teachers, src_path)

    # srcファイルとして登録
    sd['files'] = {**sd.get('files', {}), 'src': src_path}
    save_session_files(sd)

    teacher_names = sorted(set(sr['name'] for sr in survey_results))

    # セッション結果にweekly_teachersを保存（講師ピッカー用）
    res = sd.get('result', {})
    res['weekly_teachers'] = weekly_teachers
    sd['result'] = res
    save_session_result(sd)

    return jsonify({
        'ok': True,
        'teachers': teacher_names,
        'teacherCount': len(teacher_names),
        'weeks': len(weekly_teachers),
        'weeklyTeachers': _sanitize_weekly_teachers(weekly_teachers),
        'surveyNameMap': survey_name_map,
        'errors': errors,
        'files': {k: (os.path.basename(v) if isinstance(v, str) else [os.path.basename(p) for p in v]) for k, v in sd.get('files', {}).items()},
    })

@app.route('/api/resolve_name_conflict', methods=['POST'])
@login_required
def resolve_name_conflict():
    """手動追加講師名がサーベイ講師と衝突した場合、サーベイ側をリネームする"""
    sd = get_session_data()
    data = request.get_json(force=True)
    manual_name = data.get('name', '').strip()
    if not manual_name:
        return jsonify({'conflict': False})

    survey_name_map = sd.get('survey_name_map', {})
    if manual_name not in survey_name_map:
        return jsonify({'conflict': False})

    full_name = survey_name_map[manual_name]
    parts = full_name.replace('\u3000', ' ').split()
    if len(parts) < 2:
        return jsonify({'conflict': False})

    new_short = parts[1] + 'T'
    print(f"[resolve_name_conflict] 衝突: 手動「{manual_name}」⇔ サーベイ「{full_name}」→「{new_short}」", flush=True)

    # セッション内の weekly_teachers をリネーム
    res = sd.get('result', {})
    wt = res.get('weekly_teachers')
    if wt:
        for wi in range(len(wt)):
            for day in wt[wi]:
                for ts in wt[wi][day]:
                    wt[wi][day][ts] = [new_short if t == manual_name else t for t in wt[wi][day][ts]]
        res['weekly_teachers'] = wt
        sd['result'] = res
        save_session_result(sd)

    # survey_name_map も更新
    survey_name_map[new_short] = full_name
    del survey_name_map[manual_name]
    sd['survey_name_map'] = survey_name_map
    save_session_files(sd)

    # NAME_MAP も更新
    NAME_MAP[full_name] = new_short

    return jsonify({'conflict': True, 'oldName': manual_name, 'newName': new_short, 'fullName': full_name})

@app.route('/api/consolidate_booth', methods=['POST'])
@login_required
def consolidate_booth():
    """週別ブース表ファイルとメタデータファイルを個別に保存（統合はDL時に実行）"""
    sd = get_session_data()
    meta_file = request.files.get('meta')
    week_files = request.files.getlist('weeks')

    # メタデータファイルが未指定の場合、week_filesの中から自動検出を試みる
    if not meta_file or not meta_file.filename:
        detected_meta = None
        remaining_weeks = []
        for f in week_files:
            if not f.filename: continue
            # ファイル名に「出力」を含むファイルはスキップ（前回出力ファイル）
            if '出力' in os.path.basename(f.filename):
                print(f"[consolidate] 出力ファイルをスキップ（ファイル名）: {f.filename}", flush=True)
                continue
            try:
                # 一時保存して中身を確認
                temp_path = os.path.join(sd['dir'], 'tmp_detect_' + os.path.basename(f.filename))
                f.save(temp_path)
                f.stream.seek(0)
                wb = openpyxl.load_workbook(temp_path, read_only=True)
                sheet_names = wb.sheetnames
                is_output = any(sn.startswith('_schedule_data') for sn in sheet_names)
                has_meta = any(any(k in sn for k in META_KEYWORDS) for sn in sheet_names)
                wb.close()
                if is_output:
                    print(f"[consolidate] 出力ファイルをスキップ（シート構造）: {f.filename}", flush=True)
                    os.remove(temp_path)
                    continue
                if has_meta and not detected_meta:
                    detected_meta = f
                    print(f"[consolidate] メタファイルを自動検出: {f.filename}", flush=True)
                else:
                    remaining_weeks.append(f)
                os.remove(temp_path)
            except Exception as e:
                print(f"[consolidate] 自動検出中にエラー: {f.filename} {e}", flush=True)
                remaining_weeks.append(f)

        if detected_meta:
            meta_file = detected_meta
            week_files = remaining_weeks
        else:
            return jsonify({'error': 'メタデータファイルを選択するか、フォルダ内に「必要コマ数」等を含むファイルを入れてください'}), 400

    # メタデータファイルを保存
    ok, err = validate_file(meta_file)
    if not ok:
        return jsonify({'error': f'メタデータファイル: {err}'}), 400
    meta_path = os.path.join(sd['dir'], 'meta_' + os.path.basename(meta_file.filename))
    meta_file.save(meta_path)

    try:
        meta_wb = openpyxl.load_workbook(meta_path, read_only=True)
        meta_sheet_names = [sn for sn in meta_wb.sheetnames if any(k in sn for k in META_KEYWORDS)]
        meta_wb.close()
    except Exception as e:
        return jsonify({'error': f'メタデータファイルの読み込みに失敗: {e}'}), 400

    print(f"[consolidate] メタシート: {meta_sheet_names}", flush=True)

    # 週別ファイルをバリデーション・保存
    errors = []
    saved_week_paths = []
    week_count = 0
    for f in sorted(week_files, key=lambda x: x.filename):
        ok, err = validate_file(f)
        if not ok:
            errors.append(f'{f.filename}: {err}')
            continue

        week_path = os.path.join(sd['dir'], 'week_' + os.path.basename(f.filename))
        try:
            f.save(week_path)
            week_wb = openpyxl.load_workbook(week_path, read_only=True)

            # 出力ファイル（_schedule_dataシートを含む）をスキップ
            if any(sn.startswith('_schedule_data') for sn in week_wb.sheetnames):
                print(f"[consolidate] 出力ファイルをスキップ: {f.filename}", flush=True)
                week_wb.close()
                continue

            # 有効な週シートがあるかチェック
            has_valid = False
            for sn in week_wb.sheetnames:
                if 'ブース表' not in sn:
                    continue
                if week_wb[sn].sheet_state != 'visible':
                    continue
                has_valid = True
                break

            week_wb.close()
            if has_valid:
                week_count += 1
                saved_week_paths.append(week_path)
                print(f"[consolidate] 週ファイル保存: {f.filename}", flush=True)
            else:
                print(f"[consolidate] 有効な週シートなし: {f.filename}", flush=True)
        except Exception as e:
            errors.append(f'{f.filename}: {str(e)}')
            traceback.print_exc()

    if week_count == 0:
        return jsonify({'error': '有効な週シートがありません', 'details': errors}), 400

    # メタファイルと週ファイルリストを個別に登録
    sd['files'] = {**sd.get('files', {}), 'booth': meta_path, 'week_files': saved_week_paths}
    save_session_files(sd)

    # 最終シート構成を取得
    final_sheets = []
    for wp in saved_week_paths:
        try:
            wb = openpyxl.load_workbook(wp, read_only=True)
            final_sheets.extend(sn for sn in wb.sheetnames if sn not in final_sheets)
            wb.close()
        except Exception:
            pass

    return jsonify({
        'ok': True,
        'weekCount': week_count,
        'metaSheets': meta_sheet_names,
        'removedSheets': [],
        'finalSheets': final_sheets,
        'errors': errors,
        'files': {k: (os.path.basename(v) if isinstance(v, str) else [os.path.basename(p) for p in v]) for k, v in sd.get('files', {}).items()},
    })

@app.route('/api/generate', methods=['POST'])
@login_required
def generate():
    sd = get_session_data()
    files = sd.get('files',{})
    print(f"[generate] sid={sd.get('_sid','?')}, files_keys={list(files.keys())}", flush=True)
    if not all(k in files for k in ['src','booth']):
        print(f"[generate] ERROR: ファイル不足 files={files}", flush=True)
        return jsonify({'error': 'ファイルが不足しています。再度アップロードしてください。'}), 400
    # ファイルが実際に存在するか確認
    for k in ['src', 'booth']:
        if not os.path.exists(files[k]):
            print(f"[generate] ERROR: ファイルが見つかりません: {k}={files[k]}", flush=True)
            return jsonify({'error': f'{k}ファイルが見つかりません。再度アップロードしてください。'}), 400

    data = request.get_json() or {}
    office_rule = data.get('officeRule', {d: [] for d in DAYS})
    booth_pref_ui = data.get('boothPref', {})
    booth_pref_ui = {k: int(v) for k, v in booth_pref_ui.items() if v}
    manual_teachers = data.get('manualTeachers', [])

    try:
        # メタファイルから講師スキル・ブース希望・生徒データを読み込み
        booth_wb = openpyxl.load_workbook(files['booth'])
        skills = load_teacher_skills(booth_wb)
        file_booth_pref = load_booth_pref(booth_wb)
        students = load_students_from_wb(booth_wb)
        booth_wb.close()

        # ブース希望: UI設定を優先、なければファイルから読んだ値を使用
        booth_pref = {**file_booth_pref, **booth_pref_ui}

        wt = load_weekly_teachers(files['src'])
        if not wt:
            return jsonify({'error': '元シートから出勤講師データを読み取れませんでした。シートに講師データが含まれているか確認してください。'}), 400

        # 手動追加講師はブースに配置せず候補リストにのみ表示（手動D&D用）
        if manual_teachers:
            print(f"[generate] manual teachers (候補のみ): {manual_teachers}", flush=True)

        # 週ファイルリストから週数を制限
        week_file_paths = files.get('week_files', [])
        if week_file_paths:
            if len(wt) > len(week_file_paths):
                print(f"[generate] Truncating weeks from {len(wt)} to {len(week_file_paths)} (based on week files)", flush=True)
                wt = wt[:len(week_file_paths)]
            holidays = load_holidays_from_files(week_file_paths[:len(wt)])
        else:
            # 後方互換: 統合ブックが直接アップロードされた場合
            booth_wb = openpyxl.load_workbook(files['booth'])
            valid_booth_sheets = [sn for sn in booth_wb.sheetnames if not any(k in sn for k in META_KEYWORDS)]
            if len(wt) > len(valid_booth_sheets):
                print(f"[generate] Truncating weeks from {len(wt)} to {len(valid_booth_sheets)} (based on booth sheets)", flush=True)
                wt = wt[:len(valid_booth_sheets)]
            holidays = load_holidays(booth_wb, len(wt))
            booth_wb.close()

        # 週ごとの日付情報を取得（build_scheduleに渡すために先に計算）
        if week_file_paths:
            week_dates = extract_week_dates_from_files(week_file_paths[:len(wt)])
        else:
            booth_wb = openpyxl.load_workbook(files['booth'])
            week_dates = extract_week_dates(booth_wb, len(wt))
            booth_wb.close()

        total = sum(sum(s['needs'].values()) for s in students)

        # 学習済み重みをロード
        learned_weights = load_learning_weights()

        schedule, unplaced, office_teachers = build_schedule(
            students, wt, skills, office_rule, booth_pref, holidays=holidays,
            weights=learned_weights, week_dates=week_dates
        )
        placed = sum(len(b['slots']) for w in schedule for d in w.values() for bs in d.values() for b in bs)

        # 自動チェック（生成直後に実行）
        try:
            check_issues = check_all(schedule, wt, office_teachers, students, skills)
        except Exception:
            app.logger.warning(f'auto check_all failed: {traceback.format_exc()}')
            check_issues = []
        _err_count = sum(1 for ci in check_issues if ci['level'] == 'error')
        _warn_count = sum(1 for ci in check_issues if ci['level'] == 'warn')
        check_summary = {'errors': _err_count, 'warnings': _warn_count, 'issues': check_issues}

        # JSON用にtupleをlistに変換
        schedule_json = []
        for w in schedule:
            wj = {}
            for day, ds in w.items():
                dj = {}
                for ts, booths in ds.items():
                    bj = []
                    for b in booths:
                        bj.append({'teacher': b['teacher'], 'slots': [list(s) for s in b['slots']]})
                    dj[ts] = bj
                wj[day] = dj
            schedule_json.append(wj)

        # 自動生成結果のスナップショットを保存（学習用diff比較のため）
        original_schedule_json = deepcopy(schedule_json)
        original_unplaced = deepcopy(unplaced)

        sd['result'] = {
            'schedule': schedule,
            'schedule_json': schedule_json,
            'original_schedule_json': original_schedule_json,
            'original_unplaced': original_unplaced,
            'unplaced': unplaced,
            'office_teachers': office_teachers,
            'office_rule': office_rule,
            'booth_pref': booth_pref,
            'manual_teachers': manual_teachers,
            'students': students,
            'week_dates': week_dates,
            'weekly_teachers': wt,
            'skills': skills,
        }
        save_session_result(sd)

        # 生徒データJSON化（全情報含む）
        students_json = []
        for s in students:
            avail_list = sorted([list(a) for a in s['avail']]) if s.get('avail') else None
            backup_list = sorted([list(a) for a in s['backup_avail']]) if s.get('backup_avail') else None
            fixed_list = [[d, t, subj] for d, t, subj in s.get('fixed', [])]
            students_json.append({
                'grade': s['grade'], 'name': s['name'],
                'needs': s['needs'],
                'avail': avail_list,
                'backup_avail': backup_list,
                'fixed': fixed_list,
                'notes': s.get('notes', ''),
                'ng_teachers': s['ng_teachers'],
                'wish_teachers': s['wish_teachers'],
                'ng_students': s['ng_students'],
                'ng_dates': [list(d) for d in s.get('ng_dates', set())],
            })

        return jsonify({
            'placed': placed,
            'total': total,
            'schedule': schedule_json,
            'unplaced': unplaced,
            'officeTeachers': office_teachers,
            'boothPref': booth_pref,
            'students': students_json,
            'weekDates': week_dates,
            'weeklyTeachers': _sanitize_weekly_teachers(wt),
            'checkSummary': check_summary,
        })
    except Exception as e:
        app.logger.error(f'API error: {traceback.format_exc()}')
        return jsonify({'error': '内部エラーが発生しました'}), 500

def _prepare_excel(sd, progress_fn=None):
    """Excel出力ファイルを準備し、output_pathを返す"""
    res = sd.get('result', {})
    output_path = os.path.join(sd['dir'], 'output.xlsx')

    def _prog(pct, msg):
        if progress_fn:
            progress_fn(pct, msg)

    _prog(5, 'データを準備中...')

    # --- ハッシュキャッシュ: スケジュール未変更ならファイル再生成をスキップ ---
    sched_hash = hashlib.sha256(
        json.dumps(res.get('schedule', []), sort_keys=True).encode()
    ).hexdigest()
    cached_hash = sd.get('_excel_hash')
    if cached_hash == sched_hash and os.path.exists(output_path):
        _prog(100, '完了')
        return output_path

    _prog(10, 'テンプレートを読み込み中...')

    # office_teachers が不足している場合（古いバックアップ等）、デフォルト設定で補完
    ot_list = list(res.get('office_teachers', []))
    rule = res.get('office_rule') or {d: [] for d in DAYS}
    num_sched_weeks = len(res.get('schedule', []))
    while len(ot_list) < num_sched_weeks:
        if ot_list:
            ot_list.append(dict(ot_list[-1]))
        else:
            ot_list.append({d: rule[d][0] if rule.get(d) else None for d in DAYS})

    week_file_paths = sd.get('files', {}).get('week_files')
    booth_path = sd.get('files', {}).get('booth')
    write_excel(
        res['schedule'],
        res['unplaced'],
        ot_list,
        booth_path,
        output_path,
        week_file_paths=week_file_paths,
        progress_fn=progress_fn
    )
    # ハッシュをメタデータに永続化（次回リクエストでキャッシュ判定に使用）
    sd['_excel_hash'] = sched_hash
    try:
        sid = sd.get('_sid')
        if sid:
            meta = _load_meta(sid)
            if meta:
                meta['_excel_hash'] = sched_hash
                _save_meta(sid, meta)
    except Exception:
        pass
    return output_path


@app.route('/api/download')
@login_required
def download():
    sd = get_session_data()
    res = sd.get('result', {})
    if 'schedule' not in res:
        # インメモリキャッシュが消えている場合、ディスクから復元を試みる
        sid = sd.get('_sid')
        if sid:
            disk_result = _load_result_from_disk(sid)
            if disk_result and 'schedule_json' in disk_result:
                sd['result'] = {
                    'schedule_json': disk_result['schedule_json'],
                    'schedule': disk_result['schedule_json'],
                    'unplaced': disk_result.get('unplaced', []),
                    'office_teachers': disk_result.get('office_teachers', []),
                }
                save_session_result(sd)
                res = sd['result']
    if 'schedule' not in res:
        return jsonify({'error': '先にスケジュールを生成してください'}), 400
    try:
        output_path = os.path.join(sd['dir'], 'output.xlsx')
        # ハッシュキャッシュ: _prepare_excel 内で変更有無を判定
        _prepare_excel(sd)
        return send_file(output_path, as_attachment=True, download_name='時間割_出力.xlsx')
    except Exception as e:
        app.logger.error(f'API error: {traceback.format_exc()}')
        return jsonify({'error': '内部エラーが発生しました'}), 500


import queue as _queue

@app.route('/api/download_stream')
@login_required
def download_stream():
    """SSEでExcel生成の進捗を送信し、完了後にファイルをダウンロード可能にする"""
    sd = get_session_data()
    res = sd.get('result', {})
    if 'schedule' not in res:
        # インメモリキャッシュが消えている場合、ディスクから復元を試みる
        sid = sd.get('_sid')
        if sid:
            disk_result = _load_result_from_disk(sid)
            if disk_result and 'schedule_json' in disk_result:
                sd['result'] = {
                    'schedule_json': disk_result['schedule_json'],
                    'schedule': disk_result['schedule_json'],
                    'unplaced': disk_result.get('unplaced', []),
                    'office_teachers': disk_result.get('office_teachers', []),
                }
                save_session_result(sd)
                res = sd['result']
    if 'schedule' not in res:
        def err_gen():
            yield f"data: {json.dumps({'error': '先にスケジュールを生成してください'})}\n\n"
        return app.response_class(err_gen(), mimetype='text/event-stream')

    q = _queue.Queue()

    def on_progress(pct, msg):
        q.put((pct, msg))

    def work():
        try:
            _prepare_excel(sd, progress_fn=on_progress)
            q.put(None)  # sentinel: done
        except Exception as e:
            q.put(('error', str(e)))

    t = threading.Thread(target=work, daemon=True)
    t.start()

    def generate():
        start = time.time()
        while True:
            try:
                item = q.get(timeout=5)
            except _queue.Empty:
                # 5秒ごとにSSEコメントを送信して接続を維持
                if time.time() - start > 300:
                    yield f"data: {json.dumps({'error': 'タイムアウト'})}\n\n"
                    break
                yield ": keepalive\n\n"
                continue
            if item is None:
                yield f"data: {json.dumps({'progress': 100, 'step': '完了', 'ready': True})}\n\n"
                break
            if item[0] == 'error':
                yield f"data: {json.dumps({'error': item[1]})}\n\n"
                break
            pct, msg = item
            yield f"data: {json.dumps({'progress': pct, 'step': msg}, ensure_ascii=False)}\n\n"
        t.join(timeout=5)

    return app.response_class(
        generate(),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'}
    )

def _build_state_json(sd):
    """セッションデータからスケジュール全状態のJSONシリアライズ用dictを構築"""
    res = sd.get('result', {})
    schedule = res.get('schedule_json') or res.get('schedule', [])
    students_json = []
    for s in res.get('students', []):
        students_json.append({
            'grade': s.get('grade', ''), 'name': s.get('name', ''),
            'needs': s.get('needs', {}),
            'avail': sorted([list(a) for a in s['avail']]) if s.get('avail') else None,
            'backup_avail': sorted([list(a) for a in s['backup_avail']]) if s.get('backup_avail') else None,
            'fixed': [[d, t, subj] for d, t, subj in s.get('fixed', [])],
            'notes': s.get('notes', ''),
            'ng_teachers': s.get('ng_teachers', []),
            'wish_teachers': s.get('wish_teachers', []),
            'ng_students': s.get('ng_students', []),
            'ng_dates': [list(d) for d in s.get('ng_dates', set())],
        })

    wt = res.get('weekly_teachers')

    placed = sum(len(b['slots']) for w in schedule for d in w.values() for bs in d.values() for b in bs)
    total = sum(sum(s.get('needs', {}).values()) for s in res.get('students', []))

    state_json = {
        'schedule': schedule,
        'unplaced': res.get('unplaced', []),
        'officeTeachers': res.get('office_teachers', []),
        'officeRule': res.get('office_rule', {}),
        'boothPref': res.get('booth_pref', {}),
        'manualTeachers': res.get('manual_teachers', []),
        'weekDates': res.get('week_dates'),
        'students': students_json,
        'placed': placed,
        'total': total,
    }
    if wt:
        state_json['weeklyTeachers'] = _sanitize_weekly_teachers(wt)
    return state_json


@app.route('/api/download_json')
@login_required
def download_json():
    """スケジュール全状態をJSONファイルとしてダウンロード"""
    sd = get_session_data()
    res = sd.get('result', {})
    if 'schedule' not in res:
        return jsonify({'error': '先にスケジュールを生成してください'}), 400

    try:
        state_json = _build_state_json(sd)

        json_str = json.dumps(state_json, ensure_ascii=False, indent=2)
        json_path = os.path.join(sd['dir'], 'schedule_data.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            f.write(json_str)

        return send_file(json_path, as_attachment=True, download_name='schedule_data.json',
                         mimetype='application/json')
    except Exception as e:
        app.logger.error(f'API error: {traceback.format_exc()}')
        return jsonify({'error': '内部エラーが発生しました'}), 500


# ========== クラウド保存/復元 (schedule_snapshots) ==========

@app.route('/api/cloud_save', methods=['POST'])
@login_required
def cloud_save():
    """スケジュール状態をSupabaseに永続保存 (upsert)"""
    sd = get_session_data()
    res = sd.get('result', {})
    if 'schedule' not in res and 'schedule_json' not in res:
        return jsonify({'error': 'スケジュールがありません'}), 400

    try:
        data = request.get_json(silent=True) or {}
        try:
            label = _sanitize_postgrest_value(data.get('label', 'latest'), 'label')
        except ValueError as ve:
            return jsonify({'error': str(ve)}), 400

        state = _build_state_json(sd)
        week_dates = res.get('week_dates') or {}
        year = week_dates.get('year', 0)
        month = week_dates.get('month', 0)
        if not year or not month:
            year = data.get('year', _dt.datetime.now().year)
            month = data.get('month', _dt.datetime.now().month)
        try:
            year = _sanitize_postgrest_value(year, 'int')
            month = _sanitize_postgrest_value(month, 'int')
        except ValueError as ve:
            return jsonify({'error': str(ve)}), 400

        settings = {
            'officeRule': res.get('office_rule', {}),
            'boothPref': res.get('booth_pref', {}),
            'manualTeachers': res.get('manual_teachers', []),
        }

        # schedule_only=true の場合、スケジュールデータのみ上書き（自動保存用）
        schedule_only = data.get('schedule_only', False)

        sb_body_dict = {
            'year': year,
            'month': month,
            'label': label,
            'schedule_data': state,
            'updated_at': _dt.datetime.utcnow().isoformat() + 'Z',
        }

        if not schedule_only:
            # メタデータ (skills等)
            metadata = data.get('metadata')
            if metadata is None:
                raw_skills = res.get('skills', {})
                skills_json = {t: sorted(list(s)) if isinstance(s, set) else s
                              for t, s in raw_skills.items()} if raw_skills else {}
                metadata = {'skills': skills_json}
            sb_body_dict['settings_data'] = settings
            sb_body_dict['metadata'] = metadata

            # ブース表テンプレート (include_template=true の場合のみ送信)
            if data.get('include_template', False):
                b64 = _encode_booth_files(sd)
                if b64:
                    sb_body_dict['booth_template'] = b64
                    print(f"[cloud_save] booth template included ({len(b64)} chars)", flush=True)

        print(f"[cloud_save] schedule_only={schedule_only}", flush=True)

        sb_url = f"{SUPABASE_URL}/rest/v1/schedule_snapshots?on_conflict=year,month,label"
        sb_body = json.dumps(sb_body_dict, ensure_ascii=False).encode('utf-8')
        sb_hdrs = {
            'apikey': SUPABASE_SERVICE_KEY,
            'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
            'Content-Type': 'application/json',
            'Prefer': 'resolution=merge-duplicates',
        }
        sb_req = Request(sb_url, data=sb_body, headers=sb_hdrs, method='POST')
        try:
            with urlopen(sb_req, timeout=30) as resp:
                raw = resp.read().decode('utf-8')
                print(f"[cloud_save] Supabase resp: {resp.status} len={len(raw)}", flush=True)
        except HTTPError as he:
            err_body = he.read().decode('utf-8')[:500]
            print(f"[cloud_save] Supabase HTTPError: {he.code} {err_body}", flush=True)
            return jsonify({'ok': False, 'error': 'クラウド保存に失敗しました'}), 502
        except URLError as ue:
            print(f"[cloud_save] Supabase URLError: {ue}", flush=True)
            return jsonify({'ok': False, 'error': 'クラウド接続に失敗しました'}), 502

        print(f"[cloud_save] saved {year}/{month} label={label}", flush=True)
        return jsonify({'ok': True, 'year': year, 'month': month, 'label': label})
    except Exception as e:
        app.logger.error(f'API error: {traceback.format_exc()}')
        return jsonify({'error': '内部エラーが発生しました'}), 500


@app.route('/api/cloud_list')
@login_required
def cloud_list():
    """保存済みスナップショット一覧を取得 (メタデータのみ)"""
    try:
        rows = _supabase_request('GET', 'schedule_snapshots',
            'select=id,year,month,label,created_at,updated_at'
            '&order=updated_at.desc&limit=50')
        return jsonify({'ok': True, 'snapshots': rows or []})
    except Exception as e:
        app.logger.error(f'cloud_list error: {traceback.format_exc()}')
        return jsonify({'error': '内部エラーが発生しました'}), 500


@app.route('/api/cloud_load', methods=['POST'])
@login_required
def cloud_load():
    """スナップショットをセッションに復元"""
    data = request.get_json(silent=True) or {}
    try:
        snapshot_id = _sanitize_postgrest_value(data.get('id'), 'uuid')
    except ValueError as ve:
        return jsonify({'error': str(ve)}), 400

    try:
        # booth_template は大きいため、専用リクエストで取得 (timeout長め)
        sb_url = f"{SUPABASE_URL}/rest/v1/schedule_snapshots?id=eq.{snapshot_id}&select=*"
        sb_hdrs = {
            'apikey': SUPABASE_SERVICE_KEY,
            'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
            'Content-Type': 'application/json',
        }
        sb_req = Request(sb_url, headers=sb_hdrs, method='GET')
        try:
            with urlopen(sb_req, timeout=30) as resp:
                rows = json.loads(resp.read().decode('utf-8'))
        except (URLError, HTTPError) as e:
            print(f"[cloud_load] Supabase fetch error: {e}", flush=True)
            return jsonify({'error': 'クラウド接続に失敗しました'}), 502
        if not rows:
            return jsonify({'error': 'スナップショットが見つかりません'}), 404

        snap = rows[0]
        state = snap['schedule_data']
        settings = snap.get('settings_data') or {}
        metadata = snap.get('metadata') or {}
        booth_b64 = snap.get('booth_template')

        # メタデータからskillsを復元 (list→set変換)
        skills = {}
        for t, subjs in metadata.get('skills', {}).items():
            skills[t] = set(subjs) if isinstance(subjs, list) else subjs

        # セッションに復元
        sd = get_session_data()
        schedule = state.get('schedule', [])
        sd['result'] = {
            'schedule_json': schedule,
            'schedule': schedule,
            'unplaced': state.get('unplaced', []),
            'office_teachers': state.get('officeTeachers', []),
            'office_rule': settings.get('officeRule', state.get('officeRule', {})),
            'booth_pref': settings.get('boothPref', state.get('boothPref', {})),
            'manual_teachers': settings.get('manualTeachers', state.get('manualTeachers', [])),
            'students': state.get('students', []),
            'week_dates': state.get('weekDates'),
            'weekly_teachers': _sanitize_weekly_teachers(state.get('weeklyTeachers')),
            'skills': skills,
        }

        # ブース表テンプレート復元
        has_booth = False
        if booth_b64:
            restored = _restore_booth_files(booth_b64, sd['dir'])
            if restored:
                new_files = {**sd.get('files', {})}
                if 'booth' in restored:
                    new_files['booth'] = restored['booth']
                if 'week_files' in restored:
                    new_files['week_files'] = restored['week_files']
                sd['files'] = new_files
                has_booth = True

        save_session_result(sd)
        save_session_files(sd)

        # フロントエンドに返却 (generate/restore_json と同じ形式)
        return jsonify({
            'ok': True,
            'schedule': schedule,
            'unplaced': state.get('unplaced', []),
            'officeTeachers': state.get('officeTeachers', []),
            'officeRule': settings.get('officeRule', state.get('officeRule', {})),
            'boothPref': settings.get('boothPref', state.get('boothPref', {})),
            'manualTeachers': settings.get('manualTeachers', state.get('manualTeachers', [])),
            'students': state.get('students', []),
            'weekDates': state.get('weekDates'),
            'weeklyTeachers': _sanitize_weekly_teachers(state.get('weeklyTeachers')),
            'placed': state.get('placed', 0),
            'total': state.get('total', 0),
            'hasBoothTemplate': has_booth,
        })
    except Exception as e:
        app.logger.error(f'API error: {traceback.format_exc()}')
        return jsonify({'error': '内部エラーが発生しました'}), 500


@app.route('/api/cloud_delete', methods=['POST'])
@login_required
def cloud_delete():
    """スナップショットを削除"""
    data = request.get_json(silent=True) or {}
    try:
        snapshot_id = _sanitize_postgrest_value(data.get('id'), 'uuid')
    except ValueError as ve:
        return jsonify({'error': str(ve)}), 400
    try:
        _supabase_request('DELETE', 'schedule_snapshots', f'id=eq.{snapshot_id}')
        return jsonify({'ok': True})
    except Exception as e:
        app.logger.error(f'cloud_delete error: {traceback.format_exc()}')
        return jsonify({'error': '削除中にエラーが発生しました'}), 500


@app.route('/api/upload_booth_template', methods=['POST'])
@login_required
def upload_booth_template():
    """結果画面からブース表テンプレートを（再）アップロード — フォルダ一括対応"""
    sd = get_session_data()
    week_files = request.files.getlist('weeks')
    if not week_files or not any(f.filename for f in week_files):
        return jsonify({'error': 'ブース表ファイルが必要です'}), 400

    # メタデータファイルの自動検出 + 週ファイル保存 (consolidate_booth と同じロジック)
    detected_meta = None
    remaining_weeks = []
    saved_week_paths = []
    meta_path = None
    errors = []

    for f in sorted(week_files, key=lambda x: x.filename):
        if not f.filename:
            continue
        # 出力ファイルスキップ
        if '出力' in os.path.basename(f.filename):
            continue
        ok, err = validate_file(f)
        if not ok:
            errors.append(f'{f.filename}: {err}')
            continue

        temp_path = os.path.join(sd['dir'], 'bt_' + os.path.basename(f.filename))
        try:
            f.save(temp_path)
            wb = openpyxl.load_workbook(temp_path, read_only=True)
            sheet_names = wb.sheetnames

            # 出力ファイル（_schedule_data含む）はスキップ
            if any(sn.startswith('_schedule_data') for sn in sheet_names):
                wb.close()
                os.remove(temp_path)
                continue

            has_meta = any(any(k in sn for k in META_KEYWORDS) for sn in sheet_names)
            has_booth = any('ブース表' in sn and wb[sn].sheet_state == 'visible' for sn in sheet_names)
            wb.close()

            if has_meta and not detected_meta:
                detected_meta = temp_path
                meta_path = temp_path
                print(f"[upload_booth_template] meta detected: {f.filename}", flush=True)
            elif has_booth:
                saved_week_paths.append(temp_path)
                print(f"[upload_booth_template] week file: {f.filename}", flush=True)
            else:
                os.remove(temp_path)
        except Exception as e:
            errors.append(f'{f.filename}: {str(e)}')

    if not saved_week_paths and not meta_path:
        return jsonify({'error': '有効なブース表ファイルが見つかりません', 'details': errors}), 400

    # セッションに反映
    new_files = {**sd.get('files', {})}
    if meta_path:
        new_files['booth'] = meta_path
    if saved_week_paths:
        new_files['week_files'] = sorted(saved_week_paths)
    sd['files'] = new_files
    save_session_files(sd)
    count = len(saved_week_paths) + (1 if meta_path else 0)
    print(f"[upload_booth_excel] saved {count} files (meta={'yes' if meta_path else 'no'}, weeks={len(saved_week_paths)})", flush=True)

    # ======== 週ファイルから配置済みスケジュールの読み取りを試みる ========
    schedule_data = None
    has_placed = False
    try:
        # 統合ブック（メタ+週シート一体型）か、個別週ファイルかを判定
        if meta_path and not saved_week_paths:
            # メタファイルのみ（週シートも含まれている可能性）
            wb = openpyxl.load_workbook(meta_path, data_only=True)
            vis_schedule, vis_ot = parse_schedule_from_wb(wb)
            wb.close()
            if vis_schedule:
                placed_count_total = sum(
                    len(b['slots']) for w in vis_schedule
                    for d in w.values() for bs in d.values() for b in bs
                )
                schedule_data = {'schedule': vis_schedule, 'officeTeachers': vis_ot}
                has_placed = placed_count_total > 0
                print(f"[upload_booth_excel] parsed schedule from meta file: {placed_count_total} slots", flush=True)
        elif saved_week_paths:
            # 個別週ファイルを順番に読み、統合スケジュールを構築
            all_schedule = []
            all_ot = []
            total_slots = 0
            for wp in sorted(saved_week_paths):
                wb = openpyxl.load_workbook(wp, data_only=True)
                vis_schedule, vis_ot = parse_schedule_from_wb(wb)
                wb.close()
                if vis_schedule:
                    for ws_data in vis_schedule:
                        all_schedule.append(ws_data)
                    for ot_data in vis_ot:
                        all_ot.append(ot_data)
                    total_slots += sum(
                        len(b['slots']) for w in vis_schedule
                        for d in w.values() for bs in d.values() for b in bs
                    )
            if all_schedule:
                schedule_data = {'schedule': all_schedule, 'officeTeachers': all_ot}
                has_placed = total_slots > 0
                print(f"[upload_booth_excel] parsed schedule from {len(all_schedule)} week files: {total_slots} slots", flush=True)
    except Exception as e:
        print(f"[upload_booth_excel] schedule parse warning: {e}", flush=True)

    # ======== スケジュールデータをセッションに反映 ========
    schedule_response = None
    if schedule_data and has_placed:
        try:
            res = sd.get('result', {})
            schedule = schedule_data['schedule']
            office_teachers = schedule_data['officeTeachers']

            # 既存の生徒データを取得
            students = res.get('students', [])

            # placed / total / unplaced を再計算
            placed = sum(len(b['slots']) for w in schedule for d in w.values() for bs in d.values() for b in bs)
            total = sum(sum(s.get('needs', {}).values()) for s in students)
            if total == 0:
                total = placed

            # 未配置コマを再計算
            placed_map = {}
            for week in schedule:
                for day_slots in week.values():
                    for booths in day_slots.values():
                        for b in booths:
                            for slot in b.get('slots', []):
                                key = (slot[1], slot[2])
                                placed_map[key] = placed_map.get(key, 0) + 1

            unplaced = []
            for s in students:
                name = s['name']
                grade = s.get('grade', '')
                for subj, need in s.get('needs', {}).items():
                    done = placed_map.get((name, subj), 0)
                    if done < need:
                        unplaced.append({
                            'grade': grade, 'name': name, 'subject': subj,
                            'count': need - done, 'reason': ''
                        })

            # セッション結果を更新
            res['schedule_json'] = schedule
            res['schedule'] = schedule
            res['office_teachers'] = office_teachers
            res['unplaced'] = unplaced
            sd['result'] = res
            save_session_result(sd)

            schedule_response = {
                'schedule': schedule,
                'officeTeachers': office_teachers,
                'unplaced': unplaced,
                'placed': placed,
                'total': total,
                'students': students,
            }
            # weeklyTeachers
            wt = res.get('weekly_teachers')
            if wt:
                schedule_response['weeklyTeachers'] = _sanitize_weekly_teachers(wt)
            print(f"[upload_booth_excel] schedule applied: {placed}/{total} placed", flush=True)
        except Exception as e:
            print(f"[upload_booth_excel] schedule apply error: {e}", flush=True)
            schedule_response = None

    # クラウドの既存スナップショットにもテンプレートを反映
    res = sd.get('result', {})
    week_dates = res.get('week_dates') or {}
    year = week_dates.get('year', 0)
    month = week_dates.get('month', 0)
    booth_saved = False
    try:
        year = _sanitize_postgrest_value(year, 'int') if year else 0
        month = _sanitize_postgrest_value(month, 'int') if month else 0
    except ValueError:
        year, month = 0, 0
    if year and month and SUPABASE_URL and SUPABASE_SERVICE_KEY:
        b64 = _encode_booth_files(sd)
        if b64:
            try:
                _supabase_request('PATCH', 'schedule_snapshots',
                    f'year=eq.{year}&month=eq.{month}',
                    body={'booth_template': b64, 'updated_at': _dt.datetime.utcnow().isoformat() + 'Z'},
                    headers_extra={'Prefer': 'return=minimal'})
                booth_saved = True
                print(f"[upload_booth_excel] cloud updated ({len(b64)} chars)", flush=True)
            except Exception as e:
                print(f"[upload_booth_excel] cloud update failed: {e}", flush=True)

    resp = {
        'ok': True,
        'count': count,
        'meta': bool(meta_path),
        'weeks': len(saved_week_paths),
        'cloudSaved': booth_saved,
        'hasSchedule': schedule_response is not None,
    }
    if schedule_response:
        resp.update(schedule_response)
    return jsonify(resp)


def parse_schedule_from_wb(wb):
    """保存済みExcelの視覚シートからスケジュール配置を再構築する。
    Returns: (schedule, office_teachers) or (None, None) if no valid week sheets found.
    """
    meta_sheets = set(sn for sn in wb.sheetnames if any(k in sn for k in META_KEYWORDS))
    week_sheets = [sn for sn in wb.sheetnames
                   if sn not in meta_sheets and sn != '_schedule_data' and sn != '未配置コマ']
    if not week_sheets:
        return None, None

    schedule = []
    office_teachers = []
    
    # シート名に日付が含まれることが多いので、それを使ってソートしたほうが安全だが
    # write_excelでは week_sheets 順で書き込んでいるため、ここでは単純にリスト順とする

    for sn in week_sheets:
        ws = wb[sn]
        week = {}
        ot = {}
        # 教室業務（行5）
        for day, vals in DAY_COLS.items():
            bc = vals[0]
            val = ws.cell(5, bc).value
            ot[day] = str(val).strip() if val else ''
        
        # 配置データ
        for tl, (sr, nb) in LAYOUT.items():
            ts = TIME_SHORT[tl]
            
            for day, vals in DAY_COLS.items():
                if day not in week:
                    week[day] = {}
                
                lc, gc, sc, sjc = vals[1], vals[2], vals[3], vals[4]
                
                booths = []
                # 表示上のブース数 (nb) まで読むが、データ構造としては MAX_BOOTHS 分確保する
                for bi in range(MAX_BOOTHS):
                    if bi >= nb:
                        # LAYOUTで定義されたブース数を超えた分は空データ
                        booths.append({'teacher': '', 'slots': []})
                        continue

                    r1 = sr + bi * 2
                    r2 = r1 + 1
                    
                    teacher = ws.cell(r1, lc).value
                    teacher = str(teacher).strip() if teacher else ''
                    
                    slots = []
                    # 生徒1
                    g1 = ws.cell(r1, gc).value
                    s1 = ws.cell(r1, sc).value
                    j1 = ws.cell(r1, sjc).value
                    if s1:
                        slots.append([str(g1 or ''), str(s1), str(j1 or '')])
                    
                    # 生徒2
                    g2 = ws.cell(r2, gc).value
                    s2 = ws.cell(r2, sc).value
                    j2 = ws.cell(r2, sjc).value
                    if s2:
                        slots.append([str(g2 or ''), str(s2), str(j2 or '')])
                        
                    booths.append({'teacher': teacher, 'slots': slots})
                
                week[day][ts] = booths
        
        schedule.append(week)
        office_teachers.append(ot)
        
    return schedule, office_teachers

@app.route('/api/load_saved', methods=['POST'])
@login_required
def load_saved():
    """保存済みExcelからスケジュール状態を読み込む"""
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'ファイルが選択されていません'}), 400
    ok, err = validate_file(f)
    if not ok:
        return jsonify({'error': err}), 400
    sd = get_session_data()
    path = os.path.join(sd['dir'], 'saved_' + os.path.basename(f.filename))
    f.save(path)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        
        state = {'schedule': [], 'officeTeachers': [], 'weekDates': None, 'students': []}
        
        # 1. 隠しJSONシートがあれば読み込む（weekDates, students, boothPref等のため）
        if '_schedule_data' in wb.sheetnames:
            ws = wb['_schedule_data']
            chunks = []
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                if row[0] is not None:
                    chunks.append(str(row[0]))
            data_str = ''.join(chunks)
            try:
                # gzip圧縮フォーマット（新）とプレーンJSON（旧）を自動判別
                if data_str.startswith('gzip:'):
                    data_str = gzip.decompress(base64.b64decode(data_str[5:])).decode('utf-8')
                state = json.loads(data_str)
            except Exception:
                pass  # JSON破損時は無視して視覚シートに頼る

        # 2. 視覚シートからスケジュール配置を上書き/復元（こちらを正とする）
        vis_schedule, vis_ot = parse_schedule_from_wb(wb)
        if vis_schedule:
            state['schedule'] = vis_schedule
            state['officeTeachers'] = vis_ot
        elif not state.get('schedule'):
             # JSONもなく視覚シートも解析できない場合
            wb.close()
            return jsonify({'error': 'このファイルには保存済みスケジュールが含まれていません。\nブース表DLしたファイルを選択してください。'}), 400

        # ======== 生徒データを保存済みExcelのシートから再取得 ========
        # 保存済みExcelは write_excel によりブース表のシートを含んでいるため
        # 別途ブース表アップロードが不要になる。
        if not state.get('weekDates'):
            # JSONがなくてweekDatesが不明な場合は、シート名から推測する
            num_weeks = len(state.get('schedule', []))
            extracted_wd = extract_week_dates(wb, num_weeks)
            if extracted_wd:
                state['weekDates'] = extracted_wd

        wd = state.get('weekDates') or {}
        year = wd.get('year', 2026)
        month = wd.get('month', 3)

        # set→list変換ヘルパー
        def serialize_student(s):
            avail = s.get('avail')
            backup = s.get('backup_avail')
            ng_dates = s.get('ng_dates')
            fixed = s.get('fixed')
            return {
                **s,
                'avail': sorted([list(a) for a in avail]) if isinstance(avail, set) else (avail or []),
                'backup_avail': sorted([list(a) for a in backup]) if isinstance(backup, set) else (backup or []),
                'ng_dates': [list(d) for d in ng_dates] if isinstance(ng_dates, set) else (ng_dates or []),
                'fixed': [list(f) for f in fixed] if fixed and isinstance(next(iter(fixed), None), (list, tuple)) else (fixed or []),
            }

        # 3. 生徒データとブース希望を保存済みExcelから再読み込み（常に最新の状態にする）
        # 元の `state['students']` はJSONパース結果だが、Excelシートの内容を優先する
        try:
            # 年月が必要だが、weekDatesから推測またはデフォルト値
            wd = state.get('weekDates') or {}
            year = wd.get('year', 2026)
            month = wd.get('month', 3)

            # 「必要コマ数」シートがあればそこから生徒情報をフルリロード
            if '必要コマ数' in wb.sheetnames:
                fresh_students = load_students_from_wb(wb, year, month)
                state['students'] = [serialize_student(s) for s in fresh_students]
                print(f"[load_saved] Reloaded {len(fresh_students)} students from '必要コマ数' sheet", flush=True)
            
            # 「講師ブース希望」シートがあればリロード
            fresh_booth_pref = load_booth_pref(wb)
            if fresh_booth_pref:
                 state['boothPref'] = fresh_booth_pref
                 print(f"[load_saved] Reloaded boothPref from '講師ブース希望' sheet", flush=True)

        except Exception as e:
            print(f"[load_saved] Warning: Failed to reload fresh data from sheets: {e}", flush=True)
            # 失敗しても致命的ではない（JSONデータ等を使う）ので続行

        # 必要ならファイルをセッションに登録（次回以降のために）
        # ただし saved_ は一時ファイル扱いなので、src/booth としては登録しないほうが無難
        # ここでは純粋に state を返すことに集中する


        wb.close()

        # ======== スケジュールの時間キー正規化 ========
        # 視覚シートから読んだ場合は既に正規化済みだが、古いJSON由来の場合は必要
        _time_normalize = {
            '14:55': '14', '16:00': '16', '17:05': '17',
            '18:10': '18', '19:15': '19', '20:20': '20',
        }
        raw_schedule = state.get('schedule', [])
        normalized_schedule = []
        for week in raw_schedule:
            nw = {}
            for day, day_data in week.items():
                nd = {}
                for ts_key, booths in day_data.items():
                    short_key = _time_normalize.get(ts_key, ts_key)
                    nd[short_key] = booths
                nw[day] = nd
            normalized_schedule.append(nw)
        state['schedule'] = normalized_schedule

        # ======== placed / total / unplaced を最新データで再計算 ========
        schedule = state['schedule']
        fresh_students_data = state.get('students', [])

        # 実際に配置されているコマ数
        placed = sum(len(b['slots']) for w in schedule for d in w.values() for bs in d.values() for b in bs)

        # 全必要コマ数を fresh_students から計算
        total = sum(sum(s.get('needs', {}).values()) for s in fresh_students_data)
        # total が 0 の場合は placed を使用（表示上の Infinity% を防ぐ）
        if total == 0:
            total = placed

        # 未配置コマを再計算
        # schedule 内に配置されている (name, subject) の数を集計
        placed_count = {}  # {(name, subj): count}
        for week in schedule:
            for day_slots in week.values():
                for booths in day_slots.values():
                    for b in booths:
                        for slot in b.get('slots', []):
                            key = (slot[1], slot[2])  # (name, subject)
                            placed_count[key] = placed_count.get(key, 0) + 1

        unplaced = []
        for s in fresh_students_data:
            name = s['name']
            grade = s.get('grade', '')
            for subj, need in s.get('needs', {}).items():
                done = placed_count.get((name, subj), 0)
                if done < need:
                    unplaced.append({
                        'grade': grade,
                        'name': name,
                        'subject': subj,
                        'count': need - done,
                        'reason': '保存済みファイルから復元'
                    })

        state['placed'] = placed
        state['total'] = total
        state['unplaced'] = unplaced

        # 安全策: weekDatesがない場合のデフォルト
        if not state.get('weekDates'):
            state['weekDates'] = {'year': year, 'month': month, 'weeks': []}

    except json.JSONDecodeError as e:
        return jsonify({'error': f'スケジュールデータの解析に失敗: {e}'}), 500
    except Exception as e:
        app.logger.error(f'API error: {traceback.format_exc()}')
        return jsonify({'error': '内部エラーが発生しました'}), 500

    # weeklyTeachers を取得（_schedule_data JSONに含まれている場合）
    weekly_teachers = state.get('weeklyTeachers') or state.get('weekly_teachers')
    if not weekly_teachers and 'src' in sd.get('files', {}):
        try:
            weekly_teachers = load_weekly_teachers(sd['files']['src'])
        except Exception:
            pass

    # スケジュール状態をセッションに保存（update_meta等で参照するため）
    sd['result'] = {
        'schedule_json': state['schedule'],
        'schedule': state['schedule'],
        'unplaced': state.get('unplaced', []),
        'office_teachers': state.get('officeTeachers', []),
        'booth_pref': state.get('boothPref', {}),
        'students': state.get('students', []),
        'week_dates': state.get('weekDates'),
        'weekly_teachers': weekly_teachers,
    }
    save_session_result(sd)

    # 保存済みファイル自体をテンプレートとして登録（再ダウンロード用）
    # ブース表シートを含むか確認し、booth_path 分岐で使えるよう booth に登録
    new_files = {**sd.get('files', {})}
    has_booth_template = False
    try:
        check_wb = openpyxl.load_workbook(path, read_only=True)
        check_snames = check_wb.sheetnames
        check_wb.close()
        # 週シート（ブース表テンプレート）があるか判定
        exclude = set()
        for sn in check_snames:
            if any(k in sn for k in META_KEYWORDS):
                exclude.add(sn)
            elif sn.startswith('_schedule_data') or sn == '未配置コマ':
                exclude.add(sn)
        week_sheet_names = [sn for sn in check_snames if sn not in exclude]
        has_booth_template = len(week_sheet_names) > 0

        new_files['booth'] = path
        # 統合ブックは booth_path 分岐で処理するため week_files は消す
        new_files.pop('week_files', None)
    except Exception:
        new_files['booth'] = path

    sd['files'] = new_files
    save_session_files(sd)

    resp = {'ok': True, 'hasBoothTemplate': has_booth_template, **state}
    if weekly_teachers:
        resp['weeklyTeachers'] = _sanitize_weekly_teachers(weekly_teachers)
    return jsonify(resp)

# ========== メタデータ・講師回答の事後更新 API ==========
@app.route('/api/update_meta', methods=['POST'])
@login_required
def update_meta():
    """メタデータファイルを再アップロードして、既存スケジュールの未配置コマを再計算"""
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'ファイルが選択されていません'}), 400
    ok, err = validate_file(f)
    if not ok:
        return jsonify({'error': err}), 400

    sd = get_session_data()
    result = sd.get('result', {})
    schedule = result.get('schedule') or result.get('schedule_json', [])
    if not schedule:
        return jsonify({'error': '復元済みスケジュールがありません。先にスケジュールを生成または復元してください。'}), 400

    # メタファイルを保存
    path = os.path.join(sd['dir'], 'meta_update_' + os.path.basename(f.filename))
    f.save(path)

    try:
        wb = openpyxl.load_workbook(path)
        fresh_students = load_students_from_wb(wb)
        fresh_booth_pref = load_booth_pref(wb)
        wb.close()
    except Exception as e:
        return jsonify({'error': f'メタデータの読み込みに失敗: {e}'}), 500

    if not fresh_students:
        return jsonify({'error': 'メタデータから生徒情報を読み取れませんでした'}), 400

    # booth ファイル参照を更新 + ブース表シートがあればテンプレートとしても登録
    new_files = {**sd.get('files', {}), 'booth': path}
    try:
        tmpl_wb = openpyxl.load_workbook(path, read_only=True)
        tmpl_exclude = set()
        for sn in tmpl_wb.sheetnames:
            if any(k in sn for k in META_KEYWORDS) or sn.startswith('_schedule_data') or sn == '未配置コマ':
                tmpl_exclude.add(sn)
        tmpl_weeks = [sn for sn in tmpl_wb.sheetnames if sn not in tmpl_exclude]
        tmpl_wb.close()
        if tmpl_weeks:
            # ブース表シートがある → booth_path 分岐で使うため week_files を消す
            new_files.pop('week_files', None)
    except Exception:
        pass
    sd['files'] = new_files
    save_session_files(sd)

    # ======== placed / total / unplaced を最新メタデータで再計算 ========
    placed = sum(len(b['slots']) for w in schedule for d in w.values() for bs in d.values() for b in bs)
    total = sum(sum(s.get('needs', {}).values()) for s in fresh_students)
    if total == 0:
        total = placed

    placed_count = {}
    for week in schedule:
        for day_slots in week.values():
            for booths in day_slots.values():
                for b in booths:
                    for slot in b.get('slots', []):
                        key = (slot[1], slot[2])
                        placed_count[key] = placed_count.get(key, 0) + 1

    unplaced = []
    for s in fresh_students:
        name = s.get('name', '')
        grade = s.get('grade', '')
        for subj, need in s.get('needs', {}).items():
            done = placed_count.get((name, subj), 0)
            if done < need:
                unplaced.append({
                    'grade': grade, 'name': name,
                    'subject': subj, 'count': need - done,
                    'reason': 'メタデータ更新から再計算'
                })

    # students を JSON-safe に変換
    students_json = []
    for s in fresh_students:
        students_json.append({
            'grade': s.get('grade', ''), 'name': s.get('name', ''),
            'needs': s.get('needs', {}),
            'avail': sorted([list(a) for a in s['avail']]) if s.get('avail') else None,
            'backup_avail': sorted([list(a) for a in s['backup_avail']]) if s.get('backup_avail') else None,
            'fixed': [[d, t, subj] for d, t, subj in s.get('fixed', [])],
            'notes': s.get('notes', ''),
            'ng_teachers': s.get('ng_teachers', []),
            'wish_teachers': s.get('wish_teachers', []),
            'ng_students': s.get('ng_students', []),
            'ng_dates': [list(d) for d in s.get('ng_dates', set())],
        })

    # result を更新
    result['students'] = students_json
    result['placed'] = placed
    result['total'] = total
    result['unplaced'] = unplaced
    if fresh_booth_pref:
        result['booth_pref'] = fresh_booth_pref
    sd['result'] = result
    save_session_result(sd)

    print(f"[update_meta] {len(fresh_students)}名の生徒データ更新, placed={placed}, total={total}, unplaced={len(unplaced)}", flush=True)

    return jsonify({
        'ok': True,
        'placed': placed,
        'total': total,
        'unplaced': unplaced,
        'students': students_json,
        'studentCount': len(fresh_students),
        'boothPref': fresh_booth_pref or {},
    })


# ========== Schedule update API ==========
@app.route('/api/update_schedule', methods=['POST'])
@login_required
def update_schedule():
    sd = get_session_data()
    data = request.get_json()
    if not data or 'schedule' not in data:
        return jsonify({'error': 'Invalid data'}), 400

    sched_json = data['schedule']
    schedule = []
    for w in sched_json:
        wk = {}
        for day, ds in w.items():
            dk = {}
            for ts, booths in ds.items():
                bk = []
                for b in booths:
                    bk.append({'teacher': b['teacher'], 'slots': [tuple(s) for s in b['slots']]})
                dk[ts] = bk
            wk[day] = dk
        schedule.append(wk)

    res = sd.get('result', {})
    res['schedule'] = schedule
    res['schedule_json'] = sched_json
    res['unplaced'] = data.get('unplaced', [])
    if data.get('students'):
        res['students'] = data['students']
    sd['result'] = res
    save_session_result(sd)

    placed = sum(len(b['slots']) for w in schedule for d in w.values() for bs in d.values() for b in bs)
    return jsonify({'ok': True, 'placed': placed})

# ========== スケジュールチェック API ==========
def _ts_label(ts):
    """短縮時間 '16' → '16:00' 等に変換"""
    return TIME_SHORT_REV.get(ts, ts)

_BL = ['①','②','③','④','⑤','⑥']

def _loc(wi, day, ts=None, bi=None):
    """位置情報文字列を生成"""
    s = f'第{wi+1}週 {day}曜'
    if ts is not None:
        s += f' {_ts_label(ts)}'
    if bi is not None:
        s += f' ブース{_BL[bi] if bi < len(_BL) else bi+1}'
    return s

def check_all(schedule, weekly_teachers, office_teachers, students, skills):
    """全チェックを最小パス数で実行する統合チェッカー"""
    issues = []
    # ---- 事前インデックス構築 ----
    smap = {s['name']: s for s in students}
    # 生徒のavail/backup_availをsetに変換して高速参照
    avail_sets = {}   # name → set of (day, ts)
    backup_sets = {}  # name → set of (day, ts)
    ng_date_sets = {} # name → set of (wi, day)
    # NG講師/NG生徒をsetに変換して高速参照
    ng_teacher_sets = {}  # name → set
    ng_student_sets = {}  # name → set
    for s in students:
        nm = s['name']
        a = s.get('avail')
        if a is not None:
            avail_sets[nm] = {(p[0], p[1]) if isinstance(p, (list, tuple)) else p for p in a}
        b = s.get('backup_avail')
        if b is not None:
            backup_sets[nm] = {(p[0], p[1]) if isinstance(p, (list, tuple)) else p for p in b}
        nd = s.get('ng_dates', [])
        if nd:
            ng_date_sets[nm] = {(d[0], d[1]) if isinstance(d, (list, tuple)) else d for d in nd}
        ngt = s.get('ng_teachers', [])
        if ngt:
            ng_teacher_sets[nm] = set(ngt)
        ngs = s.get('ng_students', [])
        if ngs:
            ng_student_sets[nm] = set(ngs)

    # W3用: (wi, day) → {name → {subj: count}}
    day_subj_counts = {}
    # W7用: (wi, day) → {name → count}   1日のコマ数集計
    day_student_counts = {}
    # W8用: (wi, day, name) → set of ts_index   時間帯インデックス集計
    _TS_ORDER = ['14', '16', '17', '18', '19', '20']
    _TS_IDX = {ts: i for i, ts in enumerate(_TS_ORDER)}
    day_student_ts = {}
    # E4用: seen pairs
    ng_pair_seen = set()

    # ---- メインパス: schedule[wi][day][ts][bi] を1回走査 ----
    for wi, week in enumerate(schedule):
        wt = weekly_teachers[wi] if wi < len(weekly_teachers) else {}
        ot = office_teachers[wi] if wi < len(office_teachers) else {}
        for day, day_data in week.items():
            ot_teacher = ot.get(day)
            ot_active = ot_teacher and ot_teacher != '休塾日'
            # E1用: その日に出勤可能な全講師（時間帯間の補間を考慮）
            day_all_teachers = set()
            for ts_teachers in wt.get(day, {}).values():
                day_all_teachers.update(ts_teachers)
            for ts, booths in day_data.items():
                teacher_booths = {}  # E6: teacher → [bi]
                ts_idx = _TS_IDX.get(ts)
                for bi, booth in enumerate(booths):
                    t = booth.get('teacher')
                    slots = booth.get('slots', [])

                    # E1: 講師未出勤（その日のいずれかの時間帯に出勤可能か）
                    if t and t not in day_all_teachers:
                        issues.append({'level': 'error', 'code': 'E1', 'title': '講師未出勤',
                            'message': f'{_loc(wi, day, ts, bi)} — {t} はこの日に出勤していません',
                            'wi': wi, 'day': day, 'ts': ts, 'bi': bi})

                    # E6: 講師重複集計
                    if t:
                        teacher_booths.setdefault(t, []).append(bi)

                    # E7: 教室業務重複
                    if ot_active and t == ot_teacher:
                        issues.append({'level': 'error', 'code': 'E7', 'title': '教室業務重複',
                            'message': f'{_loc(wi, day, ts, bi)} — 教室業務担当 {ot_teacher} がブースにも配置されています',
                            'wi': wi, 'day': day, 'ts': ts, 'bi': bi})

                    # スロットレベルのチェック
                    names_in_booth = []
                    for slot in slots:
                        sname = slot[1] if len(slot) > 1 else None
                        if not sname:
                            continue
                        names_in_booth.append(sname)

                        # E3: NG講師
                        if t and sname in ng_teacher_sets and t in ng_teacher_sets[sname]:
                            issues.append({'level': 'error', 'code': 'E3', 'title': 'NG講師',
                                'message': f'{_loc(wi, day, ts, bi)} — {sname} のNG講師 {t} に配置されています',
                                'wi': wi, 'day': day, 'ts': ts, 'bi': bi})

                        # W1: 希望時間外
                        has_avail = sname in avail_sets
                        has_backup = sname in backup_sets
                        if has_avail or has_backup:
                            pair = (day, ts)
                            in_avail = has_avail and pair in avail_sets[sname]
                            in_backup = has_backup and pair in backup_sets[sname]
                            if not in_avail and not in_backup:
                                issues.append({'level': 'warn', 'code': 'W1', 'title': '希望時間外',
                                    'message': f'{_loc(wi, day, ts, bi)} — {sname} の希望/予備時間外に配置されています',
                                    'wi': wi, 'day': day, 'ts': ts, 'bi': bi})

                        # W2: NG日程
                        if sname in ng_date_sets and (wi, day) in ng_date_sets[sname]:
                            issues.append({'level': 'warn', 'code': 'W2', 'title': 'NG日程',
                                'message': f'{_loc(wi, day, ts, bi)} — {sname} のNG日程に配置されています',
                                'wi': wi, 'day': day, 'ts': ts, 'bi': bi})

                        # W4: 指導スキル不足
                        if skills and t and len(slot) >= 3:
                            grade, subj = slot[0], slot[2]
                            if not can_teach(t, grade, subj, skills):
                                issues.append({'level': 'warn', 'code': 'W4', 'title': '指導スキル不足',
                                    'message': f'{_loc(wi, day, ts, bi)} — {t} は {grade} {subj} を指導できません（生徒: {sname}）',
                                    'wi': wi, 'day': day, 'ts': ts, 'bi': bi})

                        # W3集計: 同日同科目
                        if len(slot) >= 3:
                            subj = slot[2]
                            dk = (wi, day)
                            dsc = day_subj_counts.setdefault(dk, {}).setdefault(sname, {})
                            dsc[subj] = dsc.get(subj, 0) + 1

                        # W7/W8集計: 1日あたりのコマ数・時間帯
                        dk2 = (wi, day)
                        dsc2 = day_student_counts.setdefault(dk2, {})
                        dsc2[sname] = dsc2.get(sname, 0) + 1
                        if ts_idx is not None:
                            day_student_ts.setdefault((wi, day, sname), set()).add(ts_idx)

                    # E4: NG生徒ペア
                    if len(names_in_booth) >= 2:
                        for i, a in enumerate(names_in_booth):
                            for b_name in names_in_booth[i+1:]:
                                is_ng = (a in ng_student_sets and b_name in ng_student_sets[a]) or \
                                        (b_name in ng_student_sets and a in ng_student_sets[b_name])
                                if is_ng:
                                    key = (wi, day, ts, bi, tuple(sorted([a, b_name])))
                                    if key not in ng_pair_seen:
                                        ng_pair_seen.add(key)
                                        issues.append({'level': 'error', 'code': 'E4', 'title': 'NG生徒',
                                            'message': f'{_loc(wi, day, ts, bi)} — {a} と {b_name} はNG生徒ペアです',
                                            'wi': wi, 'day': day, 'ts': ts, 'bi': bi})

                # E6: 講師重複判定
                for t, bis in teacher_booths.items():
                    if len(bis) > 1:
                        bl = ', '.join(_BL[b] if b < len(_BL) else str(b+1) for b in bis)
                        issues.append({'level': 'error', 'code': 'E6', 'title': '講師重複',
                            'message': f'{_loc(wi, day, ts)} — {t} がブース{bl}に重複配置されています',
                            'wi': wi, 'day': day, 'ts': ts})

    # ---- E2: 教室業務講師出勤チェック（office_teachersループ）----
    for wi, ot in enumerate(office_teachers):
        wt = weekly_teachers[wi] if wi < len(weekly_teachers) else {}
        for day, teacher in ot.items():
            if not teacher or teacher == '休塾日':
                continue
            day_data = wt.get(day, {})
            found = any(teacher in teachers for teachers in day_data.values())
            if not found:
                issues.append({'level': 'error', 'code': 'E2', 'title': '教室業務講師未出勤',
                    'message': f'{_loc(wi, day)} — 教室業務担当 {teacher} はこの日に出勤していません',
                    'wi': wi, 'day': day, 'ts': None})

    # ---- W3: 同日同科目の判定 ----
    for (wi, day), name_subjs in day_subj_counts.items():
        for sname, subjs in name_subjs.items():
            for subj, cnt in subjs.items():
                if cnt >= 2:
                    issues.append({'level': 'warn', 'code': 'W3', 'title': '同日同科目',
                        'message': f'{_loc(wi, day)} — {sname} の {subj} が同じ曜日に{cnt}回配置されています',
                        'wi': wi, 'day': day, 'ts': None})

    # ---- W7: 1日3コマ以上配置 ----
    for (wi, day), name_counts in day_student_counts.items():
        for sname, cnt in name_counts.items():
            if cnt >= 3:
                issues.append({'level': 'warn', 'code': 'W7', 'title': '1日3コマ以上',
                    'message': f'{_loc(wi, day)} — {sname} が同日に{cnt}コマ配置されています',
                    'wi': wi, 'day': day, 'ts': None})

    # ---- W8: 1コマ空き配置（連続しない時間帯） ----
    for (wi, day, sname), ts_set in day_student_ts.items():
        if len(ts_set) < 2:
            continue
        sorted_ts = sorted(ts_set)
        for i in range(len(sorted_ts) - 1):
            gap = sorted_ts[i + 1] - sorted_ts[i]
            if gap == 2:
                ts_a = _TS_ORDER[sorted_ts[i]]
                ts_b = _TS_ORDER[sorted_ts[i + 1]]
                issues.append({'level': 'warn', 'code': 'W8', 'title': '1コマ空き',
                    'message': f'{_loc(wi, day)} — {sname} が {_ts_label(ts_a)} と {_ts_label(ts_b)} に配置（1コマ空き）',
                    'wi': wi, 'day': day, 'ts': None})

    return issues

@app.route('/api/check', methods=['GET'])
@login_required
def check_schedule():
    """スケジュールの制約違反をチェックする（セッションデータを使用）"""
    sd = get_session_data()
    res = sd.get('result', {})
    schedule = res.get('schedule_json') or res.get('schedule', [])
    office_teachers = res.get('office_teachers', [])
    students = res.get('students', [])

    # weekly_teachers: セッションにあればそれを使用、なければsrcから再読み込み
    weekly_teachers = res.get('weekly_teachers')
    if not weekly_teachers and 'src' in sd.get('files', {}):
        try:
            weekly_teachers = load_weekly_teachers(sd['files']['src'])
        except Exception:
            weekly_teachers = []

    # skills: セッションにあればそれを使用、なければブース表から再読み込み
    skills = res.get('skills', {})
    if not skills:
        booth_path = sd.get('files', {}).get('booth')
        if booth_path and os.path.exists(booth_path):
            try:
                booth_wb = openpyxl.load_workbook(booth_path, data_only=True, read_only=True)
                skills = load_teacher_skills(booth_wb)
                booth_wb.close()
            except Exception:
                pass

    issues = check_all(schedule, weekly_teachers or [], office_teachers, students, skills)

    return jsonify({
        'issues': issues,
        'errorCount': sum(1 for i in issues if i['level'] == 'error'),
        'warnCount': sum(1 for i in issues if i['level'] == 'warn'),
    })

# ========== JSON restore API ==========
@app.route('/api/restore_json', methods=['POST'])
@login_required
def restore_json():
    """JSONバックアップ + 既存ファイルからスケジュール状態を復元する"""
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'ファイルが選択されていません'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext != '.json':
        return jsonify({'error': f'JSONファイルを選択してください（選択: {ext}）'}), 400

    sd = get_session_data()
    try:
        raw = f.read().decode('utf-8')
        state = json.loads(raw)
    except (UnicodeDecodeError, json.JSONDecodeError) as e:
        return jsonify({'error': f'JSONの解析に失敗しました: {e}'}), 400

    schedule = state.get('schedule') or state.get('schedule_json')
    if not schedule:
        return jsonify({'error': 'スケジュールデータが含まれていません'}), 400

    unplaced = state.get('unplaced', [])
    office_teachers = state.get('officeTeachers') or state.get('office_teachers', [])
    office_rule = state.get('officeRule') or state.get('office_rule', {})
    booth_pref = state.get('boothPref') or state.get('booth_pref', {})
    manual_teachers = state.get('manualTeachers') or state.get('manual_teachers', [])
    students = state.get('students', [])
    week_dates = state.get('weekDates') or state.get('week_dates')
    weekly_teachers = state.get('weeklyTeachers') or state.get('weekly_teachers')
    placed = state.get('placed', 0)
    total = state.get('total', 0)

    # ブース表ファイル（メタ + 週別）を処理
    files = dict(sd.get('files', {}))
    booth_files = request.files.getlist('booth_files')
    if booth_files and any(bf.filename for bf in booth_files):
        # メタファイルと週ファイルを自動検出
        detected_meta = None
        week_paths = []
        for bf in sorted(booth_files, key=lambda x: x.filename or ''):
            if not bf.filename:
                continue
            fname = os.path.basename(bf.filename)
            if '出力' in fname:
                continue
            ok, err = validate_file(bf)
            if not ok:
                continue
            temp_path = os.path.join(sd['dir'], 'tmp_' + fname)
            bf.save(temp_path)
            try:
                wb = openpyxl.load_workbook(temp_path, read_only=True)
                snames = wb.sheetnames
                is_output = any(sn.startswith('_schedule_data') for sn in snames)
                has_meta = any(any(k in sn for k in META_KEYWORDS) for sn in snames)
                wb.close()
                if is_output:
                    os.remove(temp_path)
                    continue
                if has_meta and not detected_meta:
                    meta_path = os.path.join(sd['dir'], 'meta_' + fname)
                    os.rename(temp_path, meta_path)
                    detected_meta = meta_path
                    files['booth'] = meta_path
                else:
                    week_path = os.path.join(sd['dir'], 'week_' + fname)
                    os.rename(temp_path, week_path)
                    # 有効な週シートがあるか確認
                    wb2 = openpyxl.load_workbook(week_path, read_only=True)
                    has_valid = any('ブース表' in sn and wb2[sn].sheet_state == 'visible' for sn in wb2.sheetnames)
                    wb2.close()
                    if has_valid:
                        week_paths.append(week_path)
            except Exception:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
        if week_paths:
            files['week_files'] = week_paths
        print(f"[restore_json] meta={'yes' if detected_meta else 'no'}, weeks={len(week_paths)}", flush=True)
    else:
        # 後方互換: 個別の booth / src ファイル
        for key in ['booth', 'src']:
            fx = request.files.get(key)
            if fx and fx.filename:
                ok, err = validate_file(fx)
                if not ok:
                    return jsonify({'error': f'{key}ファイル: {err}'}), 400
                path = os.path.join(sd['dir'], key + '_' + fx.filename)
                fx.save(path)
                files[key] = path
    # 講師回答ファイルを処理（任意）
    survey_files = request.files.getlist('surveys')
    survey_wt = None
    survey_teacher_count = 0
    survey_errors = []
    if survey_files and any(sf.filename for sf in survey_files):
        print(f"[restore_json] survey files: {[sf.filename for sf in survey_files]}", flush=True)
        survey_results, survey_errors, survey_name_map = _process_survey_files(survey_files, sd['dir'])
        if survey_results:
            survey_wt = aggregate_surveys_to_weekly(survey_results)
            src_path = os.path.join(sd['dir'], 'generated_src.xlsx')
            generate_src_excel(survey_wt, src_path)
            files['src'] = src_path
            sd['survey_name_map'] = survey_name_map
            survey_teacher_count = len(set(sr['name'] for sr in survey_results))
            print(f"[restore_json] survey: {survey_teacher_count}名, {len(survey_wt)}週", flush=True)

    sd['files'] = files
    save_session_files(sd)

    # srcがあればweeklyTeachersを再取得（最新化）
    if 'src' in files and not survey_wt:
        try:
            weekly_teachers = load_weekly_teachers(files['src'])
        except Exception:
            pass

    # サーベイデータとJSONデータをマージ
    if survey_wt:
        weekly_teachers = _merge_weekly_teachers(weekly_teachers, survey_wt)

    # メタデータExcelからJSONに不足しているデータを補完
    if 'booth' in files:
        try:
            meta_wb = openpyxl.load_workbook(files['booth'])
            if not students:
                students = load_students_from_wb(meta_wb)
                print(f"[restore_json] students補完: {len(students)}名", flush=True)
            if not booth_pref:
                booth_pref = load_booth_pref(meta_wb)
            meta_wb.close()
        except Exception as e:
            print(f"[restore_json] メタ補完エラー: {e}", flush=True)

    # 週ファイルからweek_datesを補完
    week_file_paths = files.get('week_files', [])
    if week_file_paths:
        if not week_dates or not week_dates.get('weeks'):
            week_dates = extract_week_dates_from_files(week_file_paths)
            print(f"[restore_json] week_dates補完: {week_dates}", flush=True)

    # ======== placed / total / unplaced を最新データで再計算 ========
    placed = sum(len(b['slots']) for w in schedule for d in w.values() for bs in d.values() for b in bs)
    total = sum(sum(s.get('needs', {}).values()) for s in students) if students else 0
    if total == 0:
        total = placed

    # 未配置コマを再計算（studentsがある場合）
    if students:
        placed_count = {}
        for week in schedule:
            for day_slots in week.values():
                for booths in day_slots.values():
                    for b in booths:
                        for slot in b.get('slots', []):
                            key = (slot[1], slot[2])
                            placed_count[key] = placed_count.get(key, 0) + 1
        unplaced = []
        for s in students:
            name = s.get('name', '') if isinstance(s, dict) else s['name']
            grade = s.get('grade', '') if isinstance(s, dict) else s.get('grade', '')
            for subj, need in (s.get('needs', {}) if isinstance(s, dict) else s.get('needs', {})).items():
                done = placed_count.get((name, subj), 0)
                if done < need:
                    unplaced.append({
                        'grade': grade, 'name': name,
                        'subject': subj, 'count': need - done,
                        'reason': 'JSON復元から再計算'
                    })
        print(f"[restore_json] 再計算: placed={placed}, total={total}, unplaced={len(unplaced)}件", flush=True)
    else:
        print(f"[restore_json] WARNING: students is empty. booth={'booth' in files}, booth_files_sent={bool(booth_files and any(bf.filename for bf in booth_files))}", flush=True)

    sd['result'] = {
        'schedule_json': schedule,
        'schedule': schedule,
        'unplaced': unplaced,
        'office_teachers': office_teachers,
        'office_rule': office_rule,
        'booth_pref': booth_pref,
        'manual_teachers': manual_teachers,
        'students': students,
        'week_dates': week_dates,
        'weekly_teachers': _sanitize_weekly_teachers(weekly_teachers),  # ダウンロード時のフォールバック用
    }
    save_session_result(sd)

    # students JSON化（setやtupleをlist化）
    students_json = []
    for s in students:
        students_json.append({
            'grade': s.get('grade', ''), 'name': s.get('name', ''),
            'needs': s.get('needs', {}),
            'avail': sorted([list(a) for a in s['avail']]) if s.get('avail') else None,
            'backup_avail': sorted([list(a) for a in s['backup_avail']]) if s.get('backup_avail') else None,
            'fixed': [[d, t, subj] for d, t, subj in s.get('fixed', [])],
            'notes': s.get('notes', ''),
            'ng_teachers': s.get('ng_teachers', []),
            'wish_teachers': s.get('wish_teachers', []),
            'ng_students': s.get('ng_students', []),
            'ng_dates': [list(d) for d in s.get('ng_dates', set())],
        })

    resp = {
        'ok': True,
        'placed': placed,
        'total': total,
        'schedule': schedule,
        'unplaced': unplaced,
        'officeTeachers': office_teachers,
        'officeRule': office_rule,
        'boothPref': booth_pref,
        'manualTeachers': manual_teachers,
        'students': students_json,
        'weekDates': week_dates,
        'hasBooth': 'booth' in files,
        'hasWeekFiles': bool(files.get('week_files')),
        'weeklyTeachers': _sanitize_weekly_teachers(weekly_teachers) or [],
        'surveyTeacherCount': survey_teacher_count,
        'surveyErrors': survey_errors,
    }
    return jsonify(resp)

# ========== State persistence API ==========
@app.route('/api/state')
@login_required
def get_state():
    """保存済みスケジュール状態を返す（ページリロード時の復元用）"""
    sd = get_session_data()
    res = sd.get('result', {})

    # インメモリキャッシュにあればそれを使う
    if res and 'schedule_json' in res:
        students_raw = res.get('students', [])
        week_dates = res.get('week_dates')
        files = sd.get('files', {})

        # students が空ならメタデータExcelから補完
        if not students_raw and 'booth' in files:
            try:
                meta_wb = openpyxl.load_workbook(files['booth'])
                students_raw = load_students_from_wb(meta_wb)
                meta_wb.close()
                res['students'] = students_raw
                save_session_result(sd)
            except Exception:
                pass
        # week_dates が空なら週ファイルから補完
        if (not week_dates or not week_dates.get('weeks')) and files.get('week_files'):
            week_dates = extract_week_dates_from_files(files['week_files'])
            res['week_dates'] = week_dates
            save_session_result(sd)

        students_json = []
        for s in students_raw:
            avail_list = sorted([list(a) for a in s['avail']]) if isinstance(s.get('avail'), set) else s.get('avail')
            backup_list = sorted([list(a) for a in s['backup_avail']]) if isinstance(s.get('backup_avail'), set) else s.get('backup_avail')
            fixed_list = [list(f) for f in s.get('fixed', [])] if s.get('fixed') else []
            ng_dates_list = [list(d) for d in s.get('ng_dates', set())] if isinstance(s.get('ng_dates'), set) else s.get('ng_dates', [])
            students_json.append({
                'grade': s['grade'], 'name': s['name'],
                'needs': s.get('needs', {}),
                'avail': avail_list,
                'backup_avail': backup_list,
                'fixed': fixed_list,
                'notes': s.get('notes', ''),
                'ng_teachers': s.get('ng_teachers', []),
                'wish_teachers': s.get('wish_teachers', []),
                'ng_students': s.get('ng_students', []),
                'ng_dates': ng_dates_list,
            })
        placed = sum(len(b['slots']) for w in res['schedule_json'] for d in w.values() for bs in d.values() for b in bs)
        total = sum(sum(s.get('needs', {}).values()) for s in students_raw)
        return jsonify({
            'has_state': True,
            'placed': placed,
            'total': total,
            'schedule': res['schedule_json'],
            'unplaced': res.get('unplaced', []),
            'officeTeachers': res.get('office_teachers', []),
            'boothPref': res.get('booth_pref', {}),
            'students': students_json,
            'weekDates': week_dates or {'year':2026, 'month':3, 'weeks':[]},
        })

    # ディスクから復元を試みる
    sid = sd.get('_sid')
    if sid:
        disk_result = _load_result_from_disk(sid)
        if disk_result and 'schedule_json' in disk_result:
            students_json = []
            for s in disk_result.get('students', []):
                students_json.append({
                    'grade': s.get('grade', ''), 'name': s.get('name', ''),
                    'needs': s.get('needs', {}),
                    'avail': s.get('avail') if s.get('avail') else [],
                    'backup_avail': s.get('backup_avail') if s.get('backup_avail') else [],
                    'fixed': s.get('fixed', []),
                    'notes': s.get('notes', ''),
                    'ng_teachers': s.get('ng_teachers', []),
                    'wish_teachers': s.get('wish_teachers', []),
                    'ng_students': s.get('ng_students', []),
                    'ng_dates': s.get('ng_dates', []),
                })
            placed = sum(len(b['slots']) for w in disk_result['schedule_json'] for d in w.values() for bs in d.values() for b in bs)
            total = sum(sum(s.get('needs', {}).values()) for s in disk_result.get('students', []))

            # インメモリキャッシュに復元
            sd['result'] = {
                'schedule_json': disk_result['schedule_json'],
                'schedule': disk_result['schedule_json'],  # JSON形式で保持
                'unplaced': disk_result.get('unplaced', []),
                'office_teachers': disk_result.get('office_teachers', []),
                'booth_pref': disk_result.get('booth_pref', {}),
                'students': disk_result.get('students', []),
                'week_dates': disk_result.get('week_dates'),
                'weekly_teachers': disk_result.get('weekly_teachers'),
                'skills': disk_result.get('skills', {}),
            }
            save_session_result(sd)

            return jsonify({
                'has_state': True,
                'placed': placed,
                'total': total,
                'schedule': disk_result['schedule_json'],
                'unplaced': disk_result.get('unplaced', []),
                'officeTeachers': disk_result.get('office_teachers', []),
                'boothPref': disk_result.get('booth_pref', {}),
                'students': students_json,
                'weekDates': disk_result.get('week_dates') or {'year':2026, 'month':3, 'weeks':[]},
            })

    # Supabaseから復元を試みる（リデプロイ後のフォールバック）
    if sid:
        supa_result = _load_result_from_supabase(sid)
        if supa_result and 'schedule_json' in supa_result:
            students_json = []
            for s in supa_result.get('students', []):
                students_json.append({
                    'grade': s.get('grade', ''), 'name': s.get('name', ''),
                    'needs': s.get('needs', {}),
                    'avail': s.get('avail') if s.get('avail') else [],
                    'backup_avail': s.get('backup_avail') if s.get('backup_avail') else [],
                    'fixed': s.get('fixed', []),
                    'notes': s.get('notes', ''),
                    'ng_teachers': s.get('ng_teachers', []),
                    'wish_teachers': s.get('wish_teachers', []),
                    'ng_students': s.get('ng_students', []),
                    'ng_dates': s.get('ng_dates', []),
                })
            placed = sum(len(b['slots']) for w in supa_result['schedule_json'] for d in w.values() for bs in d.values() for b in bs)
            total = sum(sum(s.get('needs', {}).values()) for s in supa_result.get('students', []))

            # インメモリキャッシュ + ディスクに復元
            sd['result'] = {
                'schedule_json': supa_result['schedule_json'],
                'schedule': supa_result['schedule_json'],
                'original_schedule_json': supa_result.get('original_schedule_json'),
                'original_unplaced': supa_result.get('original_unplaced'),
                'unplaced': supa_result.get('unplaced', []),
                'office_teachers': supa_result.get('office_teachers', []),
                'booth_pref': supa_result.get('booth_pref', {}),
                'students': supa_result.get('students', []),
                'week_dates': supa_result.get('week_dates'),
                'weekly_teachers': supa_result.get('weekly_teachers'),
                'skills': supa_result.get('skills', {}),
            }
            save_session_result(sd)

            return jsonify({
                'has_state': True,
                'placed': placed,
                'total': total,
                'schedule': supa_result['schedule_json'],
                'unplaced': supa_result.get('unplaced', []),
                'officeTeachers': supa_result.get('office_teachers', []),
                'boothPref': supa_result.get('booth_pref', {}),
                'students': students_json,
                'weekDates': supa_result.get('week_dates') or {'year':2026, 'month':3, 'weeks':[]},
            })

    return jsonify({'has_state': False})

# ========== 学習フィードバック API ==========
@app.route('/api/submit_feedback', methods=['POST'])
@login_required
def submit_feedback():
    """手動編集後のスケジュールと自動生成結果を比較し、学習データを更新"""
    sd = get_session_data()
    res = sd.get('result', {})
    original = res.get('original_schedule_json')
    edited = res.get('schedule_json')
    if not original or not edited:
        return jsonify({'ok': False, 'error': 'スナップショットがありません', 'changes_count': 0})

    orig_unplaced = res.get('original_unplaced', [])
    edit_unplaced = res.get('unplaced', [])

    changes = compute_schedule_diff(original, edited, orig_unplaced, edit_unplaced)
    if not changes:
        return jsonify({'ok': True, 'changes_count': 0, 'summary': {}})

    # パターン抽出 & 重み調整
    signals = extract_signals(changes, original, edited)
    current_weights = load_learning_weights()
    new_weights = adjust_weights(current_weights, signals)

    # 統計更新
    stats = load_learning_stats()
    stats['session_count'] = stats.get('session_count', 0) + 1
    stats['last_updated'] = _dt.datetime.utcnow().isoformat() + 'Z'

    # 保存
    save_learning_weights(new_weights)
    save_learning_stats(stats)

    # 変更サマリ
    summary = {}
    for ch in changes:
        t = ch['type']
        summary[t] = summary.get(t, 0) + 1

    # 編集履歴保存
    placed_before = sum(len(b.get('slots', [])) for w in original for d in w.values()
                        for bs in d.values() for b in bs)
    placed_after = sum(len(b.get('slots', [])) for w in edited for d in w.values()
                       for bs in d.values() for b in bs)
    save_edit_history({
        'total_changes': len(changes),
        'placed_before': placed_before,
        'placed_after': placed_after,
        'changes': changes,
    })

    return jsonify({
        'ok': True,
        'changes_count': len(changes),
        'summary': summary,
        'weights': new_weights,
        'session_count': stats['session_count'],
    })

@app.route('/api/learning_stats')
@login_required
def learning_stats():
    """学習状況を返す"""
    stats = load_learning_stats()
    current_weights = load_learning_weights()
    return jsonify({
        'session_count': stats.get('session_count', 0),
        'last_updated': stats.get('last_updated', ''),
        'current_weights': current_weights,
        'default_weights': DEFAULT_WEIGHTS,
    })

@app.route('/api/reset_learning', methods=['POST'])
@login_required
def reset_learning():
    """学習データをリセット"""
    save_learning_weights(dict(DEFAULT_WEIGHTS))
    save_learning_stats({'session_count': 0})
    # 履歴も全削除
    rows = _supabase_request('GET', 'schedule_edit_history', 'select=id')
    if rows:
        for r in rows:
            _supabase_request('DELETE', 'schedule_edit_history', f"id=eq.{r['id']}")
    return jsonify({'ok': True})

# ========== 起動 ==========
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"\n  Booth Schedule Generator (Cloud)")
    print(f"  http://localhost:{port}\n")
    app.run(host='0.0.0.0', port=port, debug=False)
