import sys
import os
import openpyxl
import unittest
from datetime import datetime

sys.path.append(os.getcwd())

# Import app functions
from app import load_holidays, build_schedule, DAYS, WEEKDAY_TIMES, TIME_SHORT, MAX_BOOTHS

class TestFixes(unittest.TestCase):
    def test_load_holidays_filters_hidden_sheets(self):
        print("\n--- Testing load_holidays (hidden sheet filtering) ---")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Week 1: Valid
        ws1 = wb.create_sheet("Week1")
        # Set holiday cell
        # DAY_COLS: '月':(3,4,5,6,7) -> 3 is teacher col, check row 5
        ws1.cell(5, 3).value = "休塾日" # Monday
        
        # Week 2: Hidden (Should be ignored)
        ws2 = wb.create_sheet("Week2_Hidden")
        ws2.sheet_state = 'hidden'
        ws2.cell(5, 3).value = "休塾日" 
        
        # Week 3: Valid
        ws3 = wb.create_sheet("Week3")
        ws3.cell(5, 3).value = "通常" # Not holiday

        # Metadata sheet (Should be ignored)
        ws_meta = wb.create_sheet("必要コマ数")
        
        holidays = load_holidays(wb, 2) # Load 2 weeks
        
        self.assertEqual(len(holidays), 2)
        # Week 1: Monday is holiday
        self.assertTrue(holidays[0].get('月'))
        # Week 2 (loaded from Week3 sheet): Monday is NOT holiday
        # If filtered correctly, Week3 becomes the 2nd week in result
        self.assertFalse(holidays[1].get('月'))
        print("PASS: load_holidays filtered hidden/metadata sheets correctly.")

    def test_consecutive_slots(self):
        print("\n--- Testing Consecutive Slots Logic ---")
        # Mock data: S1 needs English and Math (different subjects)
        # Verify they are placed consecutively on the same day
        students = [
            {
                'name': 'S1', 'grade': 'C1', 
                'needs': {'英': 1, '数': 1}, 
                'wish_teachers': [], 'ng_teachers': [], 'ng_students': [],
                'avail': None, 
                'fixed': [],
                'backup_avail': None
            }
        ]
        
        # Mock weekly teachers
        weekly_teachers = []
        wt_week = {}
        for day in DAYS:
            dt = {}
            for tl in WEEKDAY_TIMES:
                ts = TIME_SHORT[tl]
                dt[ts] = ['T1']
            wt_week[day] = dt
        weekly_teachers.append(wt_week)
        
        # T1 teaches both
        skills = {'T1': {'中英', '中数'}} 
        office_rule = {}
        booth_pref = {}
        
        # Run scheduler
        schedule, unplaced, _ = build_schedule(students, weekly_teachers, skills, office_rule, booth_pref, holidays=[])
        
        # Verify S1 placement
        placed = []
        for day in DAYS:
            for ts, booths in schedule[0][day].items():
                for b in booths:
                    for slot in b['slots']:
                        if slot[1] == 'S1':
                            placed.append((day, ts))
        
        print(f"S1 Placed at: {placed}")
        self.assertEqual(len(placed), 2)
        
        # Check if they are on the same day
        days = set(p[0] for p in placed)
        self.assertEqual(len(days), 1, "Should be placed on the same day because of consecutive bonus")
        
        # Check if indices difference is 1 (consecutive)
        idx_list = []
        ts_map = {v: k for k, v in TIME_SHORT.items()}
        times = WEEKDAY_TIMES
        for p in placed:
            ts = p[1]
            tl = ts_map[ts]
            idx_list.append(times.index(tl))
        
        idx_list.sort()
        diff = idx_list[1] - idx_list[0]
        self.assertEqual(diff, 1, f"Slots should be consecutive. Indices: {idx_list}")
        print("PASS: Slots are consecutive.")

if __name__ == '__main__':
    unittest.main(argv=['first-arg-is-ignored'], exit=False)
